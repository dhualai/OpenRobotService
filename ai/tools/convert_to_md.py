#!/usr/bin/env python3
"""
通用文档 → Markdown 转换工具

支持格式: .docx (pandoc), .pdf (PyMuPDF), .txt (直接复制), .xlsx (openpyxl)

用法:
    # docx → FAQ（Q&A 细粒度切分）
    python tools/convert_to_md.py "D:/OpenRobotService_Data/sources/产品手册/xxx.docx" \\
        --domain team --sub-domain usp_faq

    # pdf → 产品目录
    python tools/convert_to_md.py "D:/OpenRobotService_Data/sources/产品资料/xxx.pdf" \\
        --domain company --sub-domain product_catalog

    # txt → 直接复制
    python tools/convert_to_md.py "D:/OpenRobotService_Data/sources/结构化数据/xxx.txt" \\
        --domain team --sub-domain faq

输出位置: {kb_root}/{domain}/{sub_domain}/
图片提取到: {kb_root}/{domain}/{sub_domain}/media/
"""
import argparse
import os
import re
import shutil
import subprocess
import sys
import tempfile
from pathlib import Path
from typing import Optional

# ── 默认 kb 根目录 ──────────────────────────────────────────
_DEFAULT_KB_ROOT = (
    Path(__file__).resolve().parent.parent.parent.parent
    / "OpenRobotService_Data" / "kb"
)


def _safe_filename(name: str) -> str:
    """去除路径中的非法字符"""
    return re.sub(r'[<>:"/\\|?*]', '_', name)


def convert_docx(source: Path, output_md: Path, output_media: Path) -> bool:
    """用 pandoc 将 .docx 转为 markdown，提取图片到 output_media

    返回 True 表示成功。
    """
    _say(f"  [docx] pandoc 转换中...")
    result = subprocess.run(
        [
            "pandoc", str(source),
            "-f", "docx", "-t", "markdown",
            "--wrap=none",
            f"--extract-media={output_media}",
        ],
        capture_output=True, encoding="utf-8", errors="replace",
    )
    if result.returncode != 0:
        _say(f"  [docx] pandoc 失败: {result.stderr[:200]}", file=sys.stderr)
        return False

    md_text = result.stdout
    if not md_text.strip():
        _say("  [docx] empty document")
        return False

    # ── 后处理 ──
    # 1. 去掉 pandoc TOC 锚点
    md_text = re.sub(r'\{#.*?\}', '', md_text)
    # 2. 去重连续空行
    md_text = re.sub(r'\n{4,}', '\n\n\n', md_text)
    # 3. 扁平化 pandoc 嵌套的 media/media/ → media/
    _nested = output_media / "media"
    if _nested.exists():
        for img in _nested.iterdir():
            dest = output_media / img.name
            if not dest.exists():
                shutil.move(str(img), str(dest))
        shutil.rmtree(_nested, ignore_errors=True)
    # 4. 统一所有本地图片路径 → ./media/filename（跳过 http/https 外链）
    def _fix_img_ref(m: re.Match) -> str:
        alt = m.group(1)
        raw_path = m.group(2)
        if raw_path.startswith(("http://", "https://")):
            return m.group(0)
        fname = re.split(r'[\\/]', raw_path)[-1]
        fname = re.sub(r'\{.*$', '', fname).strip()
        fname = fname.split("?")[0]
        if not fname:
            return m.group(0)
        return f"![{alt}](./media/{fname})"

    md_text = re.sub(r'!\[([^\]]*)\]\(([^)]+)\)', _fix_img_ref, md_text)
    # 5. 二次清理：去掉残留在图片引用后面的 {width=...} 属性
    md_text = re.sub(
        r'(\./media/[^)\s]+)\s*\{[^}]*\}', r'\1', md_text,
    )
    # 6. 确保文件以 # H1 开头
    if not md_text.lstrip().startswith("#"):
        md_text = f"# {source.stem}\n\n{md_text}"
    # 7. 去掉紧跟 H1 的重复标题行
    md_text = re.sub(r'^(# .+)\n\n\1\n', r'\1\n\n', md_text)

    output_md.write_text(md_text, encoding="utf-8")
    _say(f"  [docx] ✓ {output_md} ({len(md_text)} 字符)")
    _list_media(output_media)
    return True


def convert_pdf(source: Path, output_md: Path) -> bool:
    """用 PyMuPDF 提取 PDF 文本 → markdown"""
    try:
        import fitz  # PyMuPDF
    except ImportError:
        _say("  [pdf] PyMuPDF not installed: pip install pymupdf")
        return False

    _say(f"  [pdf] 提取文本中...")
    doc = fitz.open(str(source))
    lines = [f"# {source.stem}\n"]
    for i, page in enumerate(doc):
        text = page.get_text()
        if text.strip():
            lines.append(f"## 第 {i + 1} 页\n")
            lines.append(text.strip())
            lines.append("")
    doc.close()

    md_text = "\n".join(lines)
    if len(md_text) < 50:
        _say("  [pdf] less than 50 chars extracted, may be scanned PDF")
        # 仍然输出，但提醒用户
    output_md.write_text(md_text, encoding="utf-8")
    _say(f"  [pdf] ✓ {output_md} ({len(md_text)} 字符, {len(doc)} 页)")
    return True


def convert_txt(source: Path, output_md: Path) -> bool:
    """直接复制 txt → md（带最小格式化）"""
    text = source.read_text(encoding="utf-8")
    if not text.strip():
        _say("  [txt] empty file")
        return False

    # 如果第一行不是 # 标题，用文件名生成一个
    if not text.lstrip().startswith("#"):
        text = f"# {source.stem}\n\n{text}"

    output_md.write_text(text, encoding="utf-8")
    _say(f"  [txt] ✓ {output_md} ({len(text)} 字符)")
    return True


def convert_xlsx(source: Path, output_md: Path) -> bool:
    """用 openpyxl 读取 .xlsx → markdown table"""
    try:
        import openpyxl
    except ImportError:
        _say("  [xlsx] openpyxl not installed: pip install openpyxl")
        return False

    _say(f"  [xlsx] 解析中...")
    wb = openpyxl.load_workbook(str(source), data_only=True)
    lines = [f"# {source.stem}\n"]

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        lines.append(f"## {sheet_name}\n")
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            lines.append("（空表）\n")
            continue

        # 转换为 markdown table
        for r_idx, row in enumerate(rows):
            cells = [str(c) if c is not None else "" for c in row]
            # 去除换行符，避免破坏表格
            cells = [c.replace("\n", " ").replace("|", "\\|") for c in cells]
            lines.append("| " + " | ".join(cells) + " |")
            if r_idx == 0:
                # 表头分隔行
                lines.append("| " + " | ".join(["---"] * len(cells)) + " |")
        lines.append("")

    wb.close()
    md_text = "\n".join(lines)
    output_md.write_text(md_text, encoding="utf-8")
    _say(f"  [xlsx] ✓ {output_md} ({len(md_text)} 字符, {len(wb.sheetnames)} sheets)")
    return True


def _list_media(media_dir: Path) -> None:
    """列出 media 目录中的文件"""
    if not media_dir.exists():
        return
    files = list(media_dir.iterdir())
    if files:
        _say(f"  [media] {len(files)} 个图片: {', '.join(f.name for f in files[:5])}"
              + (f" ...等" if len(files) > 5 else ""))


# ── 安全的 terminal 输出 ─────────────────────────────────────
def _say(msg: str) -> None:
    """用 ASCII-safe 方式输出（避免 Chinese Windows GBK 终端报错）"""
    try:
        print(msg)
    except UnicodeEncodeError:
        print(msg.encode("ascii", errors="replace").decode("ascii"))

CONVERTERS = {
    ".docx": convert_docx,
    ".pdf":  convert_pdf,
    ".txt":  convert_txt,
    ".xlsx": convert_xlsx,
}


def main():
    parser = argparse.ArgumentParser(
        description="通用文档 → Markdown 转换工具",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog=__doc__,
    )
    parser.add_argument("source", type=str, help="源文件路径")
    parser.add_argument(
        "--domain", type=str, required=True,
        help="目标 domain（team/company/industry/project/personal）",
    )
    parser.add_argument(
        "--sub-domain", type=str, required=True,
        help="目标 sub_domain（faq/usp_manual/cheduan_errors/...）",
    )
    parser.add_argument(
        "--kb-root", type=str, default=None,
        help=f"kb 根目录（默认: {_DEFAULT_KB_ROOT}）",
    )
    parser.add_argument(
        "--title", type=str, default=None,
        help="覆盖文档标题（默认用源文件名）",
    )
    args = parser.parse_args()

    source = Path(args.source).resolve()
    if not source.exists():
        _say(f"Error: file not found: {source}")
        sys.exit(1)

    ext = source.suffix.lower()
    if ext not in CONVERTERS:
        _say(f"Unsupported format: {ext}\nSupported: {', '.join(CONVERTERS)}")
        sys.exit(1)

    kb_root = Path(args.kb_root).resolve() if args.kb_root else _DEFAULT_KB_ROOT.resolve()
    sub_dir = kb_root / args.domain / args.sub_domain
    sub_dir.mkdir(parents=True, exist_ok=True)

    output_name = _safe_filename(source.stem) + ".md"
    output_md = sub_dir / output_name
    output_media = sub_dir / "media"

    _say(f"Source:   {source}")
    _say(f"Format:   {ext}")
    _say(f"Target:   domain={args.domain}, sub_domain={args.sub_domain}")
    _say(f"Output:   {output_md}")
    _say("")

    converter = CONVERTERS[ext]

    # docx 需要 extra arg: output_media
    if ext == ".docx":
        ok = converter(source, output_md, output_media)
    else:
        ok = converter(source, output_md)

    if ok:
        _say("\nDone. Next: python -m ai.ingestion.ingest_all")
    else:
        sys.exit(1)


if __name__ == "__main__":
    main()
