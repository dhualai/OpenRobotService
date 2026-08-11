"""附件解析器：全类型附件 → 诊断上下文

支持的附件类型：
    - 压缩包 (.zip/.tar/.tgz/.gz)         → 解压 → 遍历 → 识别内部文件 → 提取
    - 文档 (.docx/.pdf/.xlsx/.csv/.txt/.md)→ 文本提取
    - 工程文件 (.log/.json/.xml/.yaml/.yml)→ 文本提取 + 结构摘要
    - 图片 (.jpg/.jpeg/.png/.webp/.bmp)   → 提取文件名列表（OCR 后续迭代）
    - 文件夹（多文件多级目录）             → 遍历目录树 → 按类型提取

设计原则：
    - 无附件不报错，静默跳过
    - 解析失败不阻塞主流程
    - 文本截断 100KB/2000 字，防止撑爆 LLM context
    - 压缩包最多遍历 50 个文件，防止炸弹
"""

import base64
import io
import os
import tarfile
import tempfile
import zipfile
from typing import List, Optional, Tuple, Set
import httpx
from pathlib import Path

from ai.agents.AiTaskPlatform.schemas import AttachmentAnalysis
from ai.core.logging import get_logger

logger = get_logger("TASK_AGENT")

# ── 常量 ──────────────────────────────────────────────────────

_TEXT_CHUNK_LIMIT = 100_000        # 单个文本文件最大字节
_EXTRACT_SUMMARY_LIMIT = 2000     # 日志摘要最多 2000 字
_ARCHIVE_MAX_FILES = 50            # 压缩包最多遍历文件数
_ARCHIVE_INNER_FILE_MAX = 100 * 1024 * 1024  # 压缩包内非日志文件最大解压大小 100MB
_DIR_MAX_FILES = 50                # 文件夹最多遍历文件数

# 单次最多分析几张图片（控制 token/时间）
_VISION_MAX_IMAGES = 3

# MIME type 映射
_EXT_TO_MIME = {
    ".jpg": "image/jpeg", ".jpeg": "image/jpeg",
    ".png": "image/png", ".bmp": "image/bmp",
    ".gif": "image/gif", ".webp": "image/webp",
}

# ── 扩展名分类 ──────────────────────────────────────────────

_IMAGE_EXTS    = {".jpg", ".jpeg", ".png", ".bmp", ".gif", ".webp"}
_LOG_EXTS      = {".txt", ".log", ".csv"}
_DOC_EXTS      = {".docx", ".pdf", ".xlsx", ".md"}
_STRUCT_EXTS   = {".json", ".xml", ".yaml", ".yml"}
_TEXTABLE_EXTS = _LOG_EXTS | _DOC_EXTS | _STRUCT_EXTS
_ARCHIVE_EXTS  = {".zip", ".tar", ".tgz", ".gz"}


def _is_image_file(filename: str, path: str = "") -> bool:
    """判断文件是否为图片类型（保留旧接口，供 router.py upload 端点使用）。"""
    return _ext(filename, path) in _IMAGE_EXTS


# ============================================================
# 对外入口
# ============================================================

async def parse_attachments(attachments: list) -> AttachmentAnalysis:
    """解析附件列表 → 分析摘要。"""
    result = AttachmentAnalysis()
    if not attachments:
        return result

    log_texts: list[str] = []
    doc_texts: list[str] = []
    struct_texts: list[str] = []
    image_names: list[str] = []

    for att in attachments:
        if not isinstance(att, dict):
            continue
        filename = att.get("filename") or att.get("name") or ""
        path = att.get("path") or att.get("url") or ""
        if not filename and not path:
            continue

        ext = _ext(filename, path)

        # ── 压缩包 ──
        if ext in _ARCHIVE_EXTS:
            try:
                inner = await _parse_archive(att, ext)
                log_texts.extend(inner)
            except Exception:
                pass

        # ── 文件夹 ──
        elif _is_directory(path):
            try:
                dir_logs, dir_docs, dir_structs = _traverse_directory(path)
                log_texts.extend(dir_logs)
                doc_texts.extend(dir_docs)
                struct_texts.extend(dir_structs)
            except Exception:
                pass

        # ── 文档 ──
        elif ext in _DOC_EXTS:
            try:
                text = await _extract_document(att, filename, ext)
                if text:
                    doc_texts.append(f"--- {filename} ---\n{text}")
            except Exception:
                pass

        # ── 工程/结构化文件 ──
        elif ext in _STRUCT_EXTS:
            try:
                text = await _extract_structured(att, filename, ext)
                if text:
                    struct_texts.append(f"--- {filename} ---\n{text}")
            except Exception:
                pass

        # ── 日志/纯文本 ──
        elif ext in _LOG_EXTS:
            try:
                content = await _read_content(att)
                if content:
                    log_texts.append(f"--- {filename} ---\n{content}")
            except Exception:
                pass

        # ── 图片 ──
        elif ext in _IMAGE_EXTS:
            image_names.append(filename or Path(path).name)

        # ── 回放 ──
        elif _is_replay_file(filename, path):
            result.has_replay = True

        # ── 未知 → 尝试当文本读 ──
        else:
            try:
                content = await _read_content(att)
                if content:
                    log_texts.append(f"--- {filename or '未命名附件'} ---\n{content}")
            except Exception:
                pass

    # ── 汇总 ──
    combined_logs = "\n".join(log_texts) if log_texts else ""
    combined_docs = "\n".join(doc_texts) if doc_texts else ""
    combined_structs = "\n".join(struct_texts) if struct_texts else ""

    if log_texts:
        result.has_logs = True
        result.log_summary = _extract_log_errors(combined_logs)

    if doc_texts:
        result.has_logs = True  # doc 内容也算分析素材
        summary = result.log_summary or ""
        result.log_summary = _merge_summaries(
            summary,
            f"[文档内容 {len(doc_texts)} 份]\n{_truncate(combined_docs, _EXTRACT_SUMMARY_LIMIT)}"
        )

    if struct_texts:
        result.has_logs = True
        summary = result.log_summary or ""
        result.log_summary = _merge_summaries(
            summary,
            f"[结构化文件 {len(struct_texts)} 份]\n{_truncate(combined_structs, _EXTRACT_SUMMARY_LIMIT)}"
        )

    if image_names:
        result.has_screenshots = True
        suffix = f"\n\n[附图片 {len(image_names)} 张：{', '.join(image_names[:10])}{' ...' if len(image_names) > 10 else ''}]"
        result.log_summary = (result.log_summary + suffix).strip()

    return result


# ============================================================
# 类型判断
# ============================================================

def _ext(filename: str, path: str) -> str:
    """识别文件扩展名，支持多段后缀（如 .tar.gz → .tgz, .log.1 → .log）。"""
    import re
    name = (filename or path).lower()
    # tar.gz 特殊处理
    if name.endswith(".tar.gz"):
        return ".tgz"
    suffixes = Path(name).suffixes
    # 日志轮转: app.log.1 / server.log.10 / syslog.1 → .log
    if re.search(r"\.(?:log|txt|csv)\.\d+$", name):
        return "." + name.rsplit(".", 2)[-2]  # 返回 .log / .txt / .csv
    if suffixes and re.match(r"^\.\d+$", suffixes[-1]):
        # 纯数字后缀，检查前面的 body 是否包含 log 关键字
        body = Path(name).stem
        if "log" in body.lower():
            return ".log"
    # 多段后缀合并: .tar.gz
    if len(suffixes) >= 2:
        combined = "".join(suffixes[-2:])
        if combined in (".tar.gz",):
            return ".tgz"
    return suffixes[-1] if suffixes else ""


def _is_directory(path: str) -> bool:
    if not path:
        return False
    return Path(path).is_dir()


def _is_replay_file(filename: str, path: str) -> bool:
    name = (filename or path).lower()
    return "replay" in name or "回放" in name


# ============================================================
# 压缩包解析
# ============================================================

async def _parse_archive(att: dict, ext: str) -> list[str]:
    """解压压缩包 → 遍历内部文件 → 按类型提取文本。"""
    raw = await _read_bytes(att)
    if not raw:
        return []

    results: list[str] = []
    try:
        if ext == ".zip":
            results = _parse_zip_bytes(raw)
        elif ext in (".tar", ".tgz", ".gz"):
            results = _parse_tar_bytes(raw, ext)
    except Exception as e:
        logger.warning(f"压缩包解析失败: {e}")

    return results


def _parse_zip_bytes(raw: bytes) -> list[str]:
    results = []
    with zipfile.ZipFile(io.BytesIO(raw)) as zf:
        count = 0
        for info in zf.infolist():
            if count >= _ARCHIVE_MAX_FILES:
                results.append(f"--- ZIP({len(zf.namelist())}文件,仅展示前{_ARCHIVE_MAX_FILES}) ---")
                break
            if info.is_dir():
                continue
            count += 1
            # 非日志文件过大时跳过（日志文件允许任意大小，_TEXT_CHUNK_LIMIT 截断）
            if _should_skip_large(info.filename, info.file_size):
                size_mb = info.file_size / (1024 * 1024)
                results.append(f"--- {info.filename} ({size_mb:.0f}MB，过大，跳过) ---")
                continue
            results.extend(_extract_inner_file(info.filename, zf.read(info.filename)))
    return results


def _parse_tar_bytes(raw: bytes, ext: str) -> list[str]:
    results = []
    bio = io.BytesIO(raw)
    # .gz / .tgz 需要先解 gzip
    if ext in (".tgz", ".gz"):
        import gzip
        bio = io.BytesIO(gzip.decompress(raw))

    try:
        with tarfile.open(fileobj=bio, mode="r:*") as tf:
            members = tf.getmembers()
            count = 0
            for member in members:
                if count >= _ARCHIVE_MAX_FILES:
                    results.append(f"--- TAR({len(members)}文件,仅展示前{_ARCHIVE_MAX_FILES}) ---")
                    break
                count += 1
                if member.isdir():
                    continue
                if _should_skip_large(member.name, member.size):
                    size_mb = member.size / (1024 * 1024)
                    results.append(f"--- {member.name} ({size_mb:.0f}MB，过大，跳过) ---")
                    continue
                try:
                    f = tf.extractfile(member)
                    if f:
                        results.extend(_extract_inner_file(member.name, f.read()))
                except Exception:
                    pass
    except tarfile.ReadError:
        # 纯 .gz 单文件
        pass

    return results


def _should_skip_large(filename: str, size: int) -> bool:
    """非日志文件超过阈值时跳过，防止大 docx/pdf/xlsx 撑爆内存。
    日志文件不受此限制——解压后 500MB+ 正常，由 _TEXT_CHUNK_LIMIT 截断。
    """
    ext = _ext(filename, "")
    if ext in _LOG_EXTS:
        return False  # 日志文件不限大小
    return size > _ARCHIVE_INNER_FILE_MAX


def _extract_inner_file(name: str, data: bytes) -> list[str]:
    """压缩包内单文件按类型提取。"""
    fname = Path(name).name
    ext = _ext(name, "")
    if name.lower().endswith(".tar.gz"):
        return _parse_tar_bytes(data, ".tgz")

    try:
        if ext in _LOG_EXTS:
            text = data.decode("utf-8", errors="replace")[:_TEXT_CHUNK_LIMIT]
            return [f"--- {fname} ---\n{text}"]
        elif ext in _DOC_EXTS:
            text = _extract_doc_bytes(data, ext)[:_TEXT_CHUNK_LIMIT]
            return [f"--- {fname} ---\n{text}"] if text else []
        elif ext in _STRUCT_EXTS:
            text = data.decode("utf-8", errors="replace")[:_TEXT_CHUNK_LIMIT]
            label = f"--- {fname} ({ext}) ---\n"
            return [f"{label}{_summarize_structured(text, ext)}"]
        elif ext in _TEXTABLE_EXTS:
            text = data.decode("utf-8", errors="replace")[:_TEXT_CHUNK_LIMIT]
            return [f"--- {fname} ---\n{text}"]
    except Exception:
        pass
    return []


# ============================================================
# 文档提取（docx / pdf / xlsx / md / csv）
# ============================================================

async def _extract_document(att: dict, filename: str, ext: str) -> Optional[str]:
    """从附件中提取文档文本。"""
    raw = await _read_bytes(att)
    if not raw:
        return None
    return _extract_doc_bytes(raw, ext)


def _extract_doc_bytes(raw: bytes, ext: str) -> Optional[str]:
    """从 bytes 提取文档文本。"""
    if ext == ".docx":
        return _extract_docx(raw)
    elif ext == ".pdf":
        return _extract_pdf(raw)
    elif ext == ".xlsx":
        return _extract_xlsx(raw)
    elif ext == ".md":
        return raw.decode("utf-8", errors="replace")[:_TEXT_CHUNK_LIMIT]
    return None


def _extract_docx(raw: bytes) -> str:
    try:
        from docx import Document
        doc = Document(io.BytesIO(raw))
        paragraphs = []
        for para in doc.paragraphs:
            if para.text.strip():
                paragraphs.append(para.text)
        return "\n".join(paragraphs)[:_TEXT_CHUNK_LIMIT]
    except Exception as e:
        logger.warning(f"docx 提取失败: {e}")
        return f"[docx 文件，提取失败: {e}]"


def _extract_pdf(raw: bytes) -> str:
    # 优先 pdfplumber（表格感知强），回退 PyPDF2
    try:
        import pdfplumber
        text_parts = []
        with pdfplumber.open(io.BytesIO(raw)) as pdf:
            for page in pdf.pages[:20]:  # 最多 20 页
                t = page.extract_text()
                if t:
                    text_parts.append(t)
        result = "\n".join(text_parts)
        if result.strip():
            return result[:_TEXT_CHUNK_LIMIT]
    except Exception as e:
        logger.debug(f"pdfplumber 提取失败，回退 PyPDF2: {e}")

    try:
        from PyPDF2 import PdfReader
        reader = PdfReader(io.BytesIO(raw))
        text_parts = []
        for page in reader.pages[:20]:
            t = page.extract_text()
            if t:
                text_parts.append(t)
        return "\n".join(text_parts)[:_TEXT_CHUNK_LIMIT]
    except Exception as e:
        logger.warning(f"PDF 提取失败: {e}")
        return f"[PDF 文件，提取失败: {e}]"


def _extract_xlsx(raw: bytes) -> str:
    try:
        import openpyxl
        wb = openpyxl.load_workbook(io.BytesIO(raw), data_only=True)
        parts = []
        for sheet_name in wb.sheetnames[:5]:  # 最多 5 个 sheet
            ws = wb[sheet_name]
            parts.append(f"Sheet: {sheet_name}")
            rows = []
            for row in ws.iter_rows(max_row=100, values_only=True):  # 最多 100 行
                cells = [str(c) if c is not None else "" for c in row]
                if any(c.strip() for c in cells):  # 跳过全空行
                    rows.append("\t".join(cells))
            parts.append("\n".join(rows))
        return "\n".join(parts)[:_TEXT_CHUNK_LIMIT]
    except ImportError:
        # 无 openpyxl → zipfile + XML 兜底
        return _extract_xlsx_fallback(raw)
    except Exception as e:
        logger.warning(f"xlsx 提取失败: {e}")
        return _extract_xlsx_fallback(raw)


def _extract_xlsx_fallback(raw: bytes) -> str:
    """xlsx 降级：xlsx 本质是 zip，提取 sharedStrings + sheet 数据。"""
    import xml.etree.ElementTree as ET
    try:
        with zipfile.ZipFile(io.BytesIO(raw)) as zf:
            # 读取共享字符串表
            strings = []
            if "xl/sharedStrings.xml" in zf.namelist():
                tree = ET.parse(io.BytesIO(zf.read("xl/sharedStrings.xml")))
                ns = {"s": "http://schemas.openxmlformats.org/spreadsheetml/2006/main"}
                for si in tree.findall(".//s:si", ns):
                    text_parts = []
                    for t in si.iter("{http://schemas.openxmlformats.org/spreadsheetml/2006/main}t"):
                        if t.text:
                            text_parts.append(t.text)
                    strings.append("".join(text_parts))

            # 读取第一个 sheet
            sheet_files = [n for n in zf.namelist() if n.startswith("xl/worksheets/sheet")]
            parts = []
            for sf in sheet_files[:2]:
                tree = ET.parse(io.BytesIO(zf.read(sf)))
                ns = {"s": "http://schemas.openxmlformats.org/spreadsheetml/2006/main"}
                for row in tree.findall(".//s:row", ns)[:100]:
                    cells = []
                    for c in row.findall("s:c", ns):
                        v = c.find("s:v", ns)
                        if v is not None and v.text:
                            t = c.get("t", "")
                            if t == "s":  # shared string
                                idx = int(v.text)
                                cells.append(strings[idx] if idx < len(strings) else "")
                            else:
                                cells.append(v.text)
                        else:
                            cells.append("")
                    if any(c.strip() for c in cells):
                        parts.append("\t".join(cells))
            return "\n".join(parts)[:_TEXT_CHUNK_LIMIT] if parts else "[xlsx 无文本数据]"
    except Exception as e:
        logger.warning(f"xlsx fallback 提取失败: {e}")
        return f"[xlsx 文件，提取失败: {e}]"


# ============================================================
# 结构化文件提取（json / xml / yaml / yml）
# ============================================================

async def _extract_structured(att: dict, filename: str, ext: str) -> Optional[str]:
    content = await _read_content(att)
    if not content:
        return None
    return _summarize_structured(content, ext)


def _summarize_structured(text: str, ext: str) -> str:
    """结构化文件：原文 + 结构摘要。"""
    truncated = text[:_TEXT_CHUNK_LIMIT]
    summary = _make_struct_summary(text, ext)
    if summary:
        return f"[结构摘要] {summary}\n\n{truncated}"
    return truncated


def _make_struct_summary(text: str, ext: str) -> str:
    """生成结构化文件的简要摘要。"""
    if ext == ".json":
        try:
            import json
            data = json.loads(text)
            if isinstance(data, dict):
                return f"JSON 对象，顶层键: {list(data.keys())[:20]}"
            elif isinstance(data, list):
                return f"JSON 数组，共 {len(data)} 项"
        except Exception:
            return ""
    elif ext in (".yaml", ".yml"):
        try:
            import yaml
            data = yaml.safe_load(text)
            if isinstance(data, dict):
                return f"YAML 对象，顶层键: {list(data.keys())[:20]}"
            elif isinstance(data, list):
                return f"YAML 数组，共 {len(data)} 项"
        except Exception:
            return ""
    elif ext == ".xml":
        import re
        tags = re.findall(r"<(\w+)[\s/>]", text[:5000])
        if tags:
            from collections import Counter
            top = Counter(tags).most_common(5)
            return f"XML，主要标签: {', '.join(f'{t}({c})' for t, c in top)}"
        return ""
    return ""


# ============================================================
# 文件夹遍历
# ============================================================

def _traverse_directory(dir_path: str) -> Tuple[list[str], list[str], list[str]]:
    """遍历本地文件夹 → 按类型分类提取。"""
    log_texts, doc_texts, struct_texts = [], [], []
    root = Path(dir_path)
    if not root.is_dir():
        return log_texts, doc_texts, struct_texts

    count = 0
    for entry in root.rglob("*"):
        if count >= _DIR_MAX_FILES:
            break
        if not entry.is_file():
            continue
        count += 1

        ext = _ext(entry.name, "")
        label = f"--- {entry.relative_to(root)} ---"
        try:
            if ext in _LOG_EXTS:
                content = entry.read_text(encoding="utf-8", errors="replace")[:_TEXT_CHUNK_LIMIT]
                log_texts.append(f"{label}\n{content}")
            elif ext in _DOC_EXTS:
                raw = entry.read_bytes()
                text = _extract_doc_bytes(raw, ext)
                if text:
                    doc_texts.append(f"{label}\n{text[:_TEXT_CHUNK_LIMIT]}")
            elif ext in _STRUCT_EXTS:
                content = entry.read_text(encoding="utf-8", errors="replace")
                struct_texts.append(f"{label}\n{_summarize_structured(content, ext)}")
            elif ext in _TEXTABLE_EXTS:
                content = entry.read_text(encoding="utf-8", errors="replace")[:_TEXT_CHUNK_LIMIT]
                log_texts.append(f"{label}\n{content}")
        except Exception:
            pass

    return log_texts, doc_texts, struct_texts


# ============================================================
# 文本内容读取
# ============================================================

async def _read_content(att: dict) -> str:
    """读取附件为文本字符串。"""
    path = att.get("path") or att.get("url", "")
    if not path:
        return ""
    path = path.replace("\\", "/")

    try:
        if path.startswith("http://") or path.startswith("https://"):
            async with httpx.AsyncClient(timeout=httpx.Timeout(10.0)) as client:
                resp = await client.get(path)
                if resp.status_code == 200:
                    return resp.text[:_TEXT_CHUNK_LIMIT]

        local = Path(path)
        if local.is_absolute() and local.exists():
            return local.read_text(encoding="utf-8", errors="replace")[:_TEXT_CHUNK_LIMIT]

        project_local = Path(__file__).resolve().parent.parent.parent / path
        if project_local.exists():
            return project_local.read_text(encoding="utf-8", errors="replace")[:_TEXT_CHUNK_LIMIT]
    except Exception as e:
        logger.warning(f"读取附件内容失败 {path}: {e}")

    return ""


async def _read_bytes(att: dict) -> Optional[bytes]:
    """读取附件为二进制。"""
    path = att.get("path") or att.get("url", "")
    if not path:
        return None
    path = path.replace("\\", "/")

    try:
        if path.startswith("http://") or path.startswith("https://"):
            async with httpx.AsyncClient(timeout=httpx.Timeout(10.0)) as client:
                resp = await client.get(path)
                if resp.status_code == 200:
                    return resp.content

        local = Path(path)
        if local.is_absolute() and local.exists():
            return local.read_bytes()

        project_local = Path(__file__).resolve().parent.parent.parent / path
        if project_local.exists():
            return project_local.read_bytes()

        # MinIO 对象路径
        if "/" in path:
            try:
                from ai.core.minio_client import minio_client
                bucket = path.split("/")[0]
                object_name = "/".join(path.split("/")[1:])
                data = minio_client.client.get_object(bucket, object_name)
                return data.read()
            except Exception as e:
                logger.warning(f"MinIO 读取失败 {path}: {e}")
    except Exception as e:
        logger.warning(f"读取附件 bytes 失败 {path}: {e}")

    return None


# ============================================================
# 日志 ERROR/WARN 提取
# ============================================================

def _extract_log_errors(text: str) -> str:
    """从日志文本提取 ERROR/WARN/异常行 + 时间线上下文。"""
    lines = text.split("\n")
    error_kws = ("ERROR", "WARN", "EXCEPTION", "FAIL", "FATAL", "Traceback")
    error_lines = [l.strip()[:200] for l in lines if any(kw in l.upper() for kw in error_kws)]

    first_ts = next((l for l in lines if _has_timestamp(l)), "")
    last_ts = next((l for l in reversed(lines) if _has_timestamp(l)), "")

    if not error_lines:
        return (
            f"日志 {len(lines)} 行，无明显错误。"
            + (f" 时间范围: {first_ts.strip()[:80]} ~ {last_ts.strip()[:80]}"
               if first_ts or last_ts else "")
        )

    parts = [f"日志 {len(lines)} 行，提取到 {len(error_lines)} 条异常：", *error_lines[:20]]
    if first_ts or last_ts:
        parts.append(f"时间范围: {first_ts.strip()[:60]} ~ {last_ts.strip()[:60]}")
    return "\n".join(parts)[:_EXTRACT_SUMMARY_LIMIT]


def _has_timestamp(line: str) -> bool:
    import re
    return bool(re.search(r"\d{2}:\d{2}:\d{2}", line))


# ============================================================
# 图片分析（视觉 LLM）
# ============================================================

async def analyze_images(attachments: list, task_context: dict = None) -> str:
    """两阶段图片分析：VLM 看图描述 → 文本模型推理分析"""
    tc = task_context or {}

    image_atts = []
    for att in attachments:
        if not isinstance(att, dict):
            continue
        filename = att.get("filename") or att.get("name") or ""
        ext = _ext(filename, att.get("path") or att.get("url", ""))
        if ext in _IMAGE_EXTS:
            image_atts.append(att)

    if not image_atts:
        return ""

    data_uris = []
    for att in image_atts[:_VISION_MAX_IMAGES]:
        fname = att.get("filename") or att.get("name") or ""
        try:
            img_bytes = await _read_bytes(att)
            if img_bytes and len(img_bytes) < 10 * 1024 * 1024:
                ext = Path(fname).suffix.lower()
                mime = _EXT_TO_MIME.get(ext, "image/png")
                b64 = base64.b64encode(img_bytes).decode()
                data_uris.append((fname, f"data:{mime};base64,{b64}"))
        except Exception:
            logger.warning(f"Failed to read image: {fname}")

    if not data_uris:
        return ""

    # Stage 1: VLM 看图描述
    try:
        from ai.core import get_llm_client
        llm = await get_llm_client()
        names = ", ".join(f"{n}" for n, _ in data_uris)
        images = [uri for _, uri in data_uris]
        desc_prompt = _build_vision_prompt(names, tc)
        raw_descriptions = await llm.complete_vision(
            prompt=desc_prompt, images=images,
            system_prompt="你是 AGV/AMR 调度系统的操作专家。仔细看图，客观描述画面内容、关键数据、异常信号和人工标注。不诊断，不下结论。",
            max_tokens=3072, temperature=0.2,
        )
    except Exception as e:
        logger.error(f"Stage 1 VLM failed: {e}")
        return f"[图片分析失败（VLM）: {e}]"

    # Stage 2: 文本模型推理
    try:
        analysis_prompt = _build_analysis_prompt(names, tc, raw_descriptions)
        analysis = await llm.complete(
            prompt=analysis_prompt,
            system_prompt="你是 AGV/AMR 领域的技术支持专家。基于图片描述和工单背景，分析线索、关联问题、提取关键发现。简洁直接。",
            max_tokens=400, temperature=0.3,
        )
    except Exception as e:
        logger.warning(f"Stage 2 text analysis failed, fallback to VLM only: {e}")
        return f"[图片分析 {len(data_uris)} 张：{names}]\n{raw_descriptions.strip()}"

    return (
        f"[图片分析 {len(data_uris)} 张：{names}]\n"
        f"📷 画面描述：\n{raw_descriptions.strip()}\n\n"
        f"🔍 线索分析：\n{analysis.strip()}"
    )


def _build_vision_prompt(names: str, task_context: dict = None) -> str:
    tc = task_context or {}
    ctx_parts = []
    if tc.get("title"): ctx_parts.append(f"标题：{tc['title']}")
    if tc.get("problem_summary"): ctx_parts.append(f"问题概述：{tc['problem_summary']}")
    elif tc.get("description"): ctx_parts.append(f"描述：{tc['description'][:150]}")
    if tc.get("hypotheses"): ctx_parts.append(f"推测方向：{'/'.join(tc['hypotheses'])}")
    if tc.get("fault_code"): ctx_parts.append(f"故障码：{tc['fault_code']}")
    if tc.get("robot_type"): ctx_parts.append(f"车型：{tc['robot_type']}")
    ctx_text = "\n".join(ctx_parts) if ctx_parts else "（无工单背景信息）"
    return (
        f"工单背景：\n{ctx_text}\n\n"
        f"现在看图片 {names}。这是 AGV/AMR 调度系统的操作截图。"
        "请仔细观察，按以下要点描述（≥300字）：\n\n"
        "**1. 画面内容** — 这是什么界面？\n"
        "**2. 关键数据** — 任务编号、车辆编号、地图名称、状态码、时间戳、坐标等字段值\n"
        "**3. 异常信号** — 红色/橙色高亮、报错弹窗、状态异常（离线/故障/取消/超时）\n"
        "**4. 人工标注** — 红色框、箭头、圈注、画笔标记、文字批注及其指向位置\n\n"
        "客观描述，不诊断，不下结论。"
    )


def _build_analysis_prompt(names: str, task_context: dict, descriptions: str) -> str:
    tc = task_context or {}
    ctx_parts = []
    if tc.get("title"): ctx_parts.append(f"工单：{tc['title']}")
    if tc.get("description"): ctx_parts.append(f"描述：{tc['description'][:150]}")
    if tc.get("problem_summary"): ctx_parts.append(f"诊断概述：{tc['problem_summary']}")
    if tc.get("hypotheses"): ctx_parts.append(f"推测方向：{'/'.join(tc['hypotheses'])}")
    ctx_text = "\n".join(ctx_parts) if ctx_parts else "（无）"
    return (
        f"工单背景：\n{ctx_text}\n\n"
        f"以下是 {names} 的画面描述：\n\n{descriptions}\n\n"
        "请基于以上描述，分析：\n"
        "1. 这些截图和工单问题有什么关联？\n"
        "2. 图中有什么值得关注的线索（异常值、状态矛盾、人工标注指向）？\n"
        "3. 下一步应该重点排查什么方向？\n\n"
        "简洁直接，工程师口吻，≤300字。"
    )


# ============================================================
# 辅助
# ============================================================

def _truncate(text: str, limit: int) -> str:
    return text if len(text) <= limit else text[:limit] + "…"


def _merge_summaries(base: str, addition: str) -> str:
    if not base:
        return addition
    if not addition:
        return base
    return f"{base}\n\n{addition}"
