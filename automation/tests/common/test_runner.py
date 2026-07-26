"""Data-driven API test runner.

Loads test cases from Excel and runs them via parametrize.

Usage:
    from automation.api.tests.test_runner import load_cases, run_case
"""

import json
from pathlib import Path

from automation.infrastructure.assertions import assert_status_code

_EXCEL_PATH = Path(__file__).parents[2] / "testdata" / "api" / "api-test-cases.xlsx"
_CACHE: dict = {}


def load_cases(module_name: str, sheet_name: str = None) -> list:
    """Load test cases for a module from Excel, cached after first load."""
    key = f"{module_name}:{sheet_name or module_name}"
    if key in _CACHE:
        return _CACHE[key]

    import openpyxl
    wb = openpyxl.load_workbook(_EXCEL_PATH, read_only=True, data_only=True)
    ws_name = sheet_name or module_name
    if ws_name not in wb.sheetnames:
        wb.close()
        return []
    ws = wb[ws_name]

    headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    cases = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] is None:
            continue
        case = dict(zip(headers, row))
        # Parse JSON strings
        for field in ("payload", "expected_fields"):
            val = case.get(field)
            if isinstance(val, str) and val.strip():
                try:
                    case[field] = json.loads(val)
                except json.JSONDecodeError:
                    pass
            elif val is None:
                case[field] = {} if field == "payload" else None
        # Normalize field names
        case["auth"] = str(case.get("auth", "N")).strip().upper()[:1]
        case["expected_status"] = int(case.get("expected_status", 200))
        cases.append(case)

    wb.close()
    _CACHE[key] = cases
    return cases


async def run_case(client, auth_header, case):
    """Execute a single test case.

    Args:
        client: httpx.AsyncClient or similar with .request()
        auth_header: dict with Authorization header
        case: dict with method, path, payload, expected_status, expected_fields
    """
    headers = auth_header if case.get("auth") == "Y" else {}
    payload = case.get("payload") or {}
    method = case["method"].upper()
    path = case["path"]

    r = await client.request(method, path, headers=headers, json=payload)
    assert_status_code(r, case.get("expected_status", 200))

    expected = case.get("expected_fields")
    if expected:
        data = r.json()
        for key, val in expected.items():
            assert data.get(key) == val, \
                f"Field {key!r}: expected {val!r}, got {data.get(key)!r}"
