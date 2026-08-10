"""Tests for cases -> Excel/Markdown export."""

import json
from pathlib import Path

from automation.ci_ai_gen.export_xlsx import (
    EXCEL_HEADERS,
    case_steps_for_execution,
    case_to_row,
    cases_from_file,
    export_cases_to_markdown,
    export_cases_to_markdown_split,
    export_cases_to_xlsx,
    group_cases_by_module,
    module_bucket,
)

CASES = [
    {
        "id": "TC001", "req_id": "REQ-01", "module": "tasks", "title": "获取任务",
        "type": "positive", "precondition": "已登录",
        "method": "GET", "path": "/api/v1/tasks/{task_id}",
        "steps": [{"id": 1, "step": "调用接口", "testData": '{"task_id": 1}', "expectedResult": "返回200"}],
    },
    {
        "id": "TC002", "req_id": "REQ-02", "module": "tasks", "title": "无权限创建任务",
        "type": "auth", "precondition": "customer 角色",
        "method": "POST", "path": "/api/v1/tasks",
        "steps": [{"id": 1, "step": "空参数", "testData": "{}", "expectedResult": "期望403"}],
    },
]


class TestCaseToRow:
    def test_headers_alignment(self):
        row = case_to_row(CASES[0])
        assert len(row) == len(EXCEL_HEADERS)

    def test_status_extracted_from_expected_result(self):
        row = case_to_row(CASES[0])
        assert row[EXCEL_HEADERS.index("expected_status")] == 200

    def test_status_403_extracted(self):
        row = case_to_row(CASES[1])
        assert row[EXCEL_HEADERS.index("expected_status")] == 403

    def test_payload_from_test_data(self):
        row = case_to_row(CASES[0])
        payload = row[EXCEL_HEADERS.index("payload")]
        assert json.loads(payload) == {"task_id": 1}

    def test_note_contains_req_and_title(self):
        row = case_to_row(CASES[0])
        note = row[EXCEL_HEADERS.index("note")]
        assert "REQ-01" in note and "获取任务" in note

    def test_auth_default_y(self):
        row = case_to_row(CASES[0])
        assert row[EXCEL_HEADERS.index("auth")] == "Y"


FLOW_CASE = {
    "id": "TC100", "req_id": "REQ-27", "module": "系统任务-工单状态流转",
    "title": "状态流转：已取消工单不可关闭", "type": "flow", "precondition": "已登录管理员",
    "method": "POST", "path": "/api/tasks",
    "steps": [
        {"id": 1, "step": "创建工单", "method": "POST", "path": "/api/tasks",
         "testData": '{"title":"流转测试单","description":"flow"}', "expectedResult": "200，返回工单含 id"},
        {"id": 2, "step": "取消工单", "method": "PATCH", "path": "/api/tasks/{{step1.body.id}}/status",
         "testData": '{"status":"cancelled"}', "expectedResult": "200"},
        {"id": 3, "step": "尝试关闭", "method": "PATCH", "path": "/api/tasks/{{step1.body.id}}/status",
         "testData": '{"status":"closed"}', "expectedResult": "400"},
    ],
}


class TestFlowStepsExport:
    def test_flow_case_exports_steps_column(self):
        row = case_to_row(FLOW_CASE)
        steps = json.loads(row[EXCEL_HEADERS.index("steps")])
        assert len(steps) == 3
        assert steps[0]["method"] == "POST"
        assert steps[0]["path"] == "/api/tasks"
        assert json.loads(steps[0]["payload"]) == {"title": "流转测试单", "description": "flow"}
        assert steps[0]["expected_status"] == 200
        assert steps[1]["path"] == "/api/tasks/{{step1.body.id}}/status"
        assert steps[2]["expected_status"] == 400

    def test_plain_case_steps_column_empty(self):
        row = case_to_row(CASES[0])
        assert row[EXCEL_HEADERS.index("steps")] == ""

    def test_case_steps_for_execution_drops_non_request_steps(self):
        case = dict(FLOW_CASE)
        case["steps"] = [
            {"id": 1, "step": "无请求的说明步骤", "testData": "x", "expectedResult": "y"},
            {"id": 2, "step": "真实请求", "method": "GET", "path": "/api/tasks/1",
             "testData": "", "expectedResult": "200"},
        ]
        exec_steps = case_steps_for_execution(case)
        assert len(exec_steps) == 1
        assert exec_steps[0]["method"] == "GET"

    def test_headers_include_steps_last(self):
        assert EXCEL_HEADERS[-1] == "steps"


class TestExportXlsx:
    def test_export_roundtrip(self, tmp_path: Path):
        out = tmp_path / "cases.xlsx"
        count = export_cases_to_xlsx(CASES, out)
        assert count == 2
        assert out.exists()

        from openpyxl import load_workbook

        wb = load_workbook(out)
        ws = wb.active
        headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
        assert headers == EXCEL_HEADERS
        rows = list(ws.iter_rows(min_row=2, values_only=True))
        assert len(rows) == 2
        assert rows[0][0] == "TC001"

    def test_cases_from_file(self, tmp_path: Path):
        p = tmp_path / "cases.json"
        p.write_text(json.dumps(CASES), encoding="utf-8")
        assert len(cases_from_file(p)) == 2


class TestExportMarkdown:
    def test_export_contains_readable_fields(self, tmp_path: Path):
        out = tmp_path / "cases.md"
        count = export_cases_to_markdown(CASES, out, title="AI 生成测试用例")
        assert count == 2
        text = out.read_text(encoding="utf-8")
        assert "共 2 条用例" in text
        assert "## TC001 获取任务" in text
        assert "**前置条件**：已登录" in text
        assert "**请求**：GET ` /api/v1/tasks/{task_id}`" in text or "GET `" in text
        assert "**操作**：调用接口" in text
        assert "测试数据" in text and "预期结果" in text
        assert "需求**：REQ-01" in text
        assert "类型**：权限" in text  # TC002 type=auth label

    def test_markdown_roundtrip_unicode(self, tmp_path: Path):
        case = dict(CASES[0])
        case["title"] = "中文标题：创建任务"
        out = tmp_path / "cases.md"
        export_cases_to_markdown([case], out)
        text = out.read_text(encoding="utf-8")
        assert "中文标题：创建任务" in text


class TestModuleSplit:
    def test_module_bucket_call(self):
        assert module_bucket("我要摇人-转工单") == "call"
        assert module_bucket("AI对话") == "call"
        assert module_bucket("认证") == "call"

    def test_module_bucket_tasks(self):
        assert module_bucket("系统任务-工单状态流转") == "tasks"
        assert module_bucket("工单管理") == "tasks"
        assert module_bucket("智能派单") == "tasks"

    def test_module_bucket_admin(self):
        assert module_bucket("后台管理-仪表盘") == "admin"
        assert module_bucket("用户管理") == "admin"
        assert module_bucket("日报管理") == "admin"

    def test_module_bucket_other(self):
        assert module_bucket("未知模块") == "other"
        assert module_bucket("") == "other"

    def test_group_cases_by_module(self):
        cases = [
            {"module": "我要摇人-转工单", "id": "TC001"},
            {"module": "系统任务-讨论区", "id": "TC002"},
            {"module": "用户管理", "id": "TC003"},
            {"module": "未知", "id": "TC004"},
        ]
        groups = group_cases_by_module(cases)
        assert [c["id"] for c in groups["call"]] == ["TC001"]
        assert [c["id"] for c in groups["tasks"]] == ["TC002"]
        assert [c["id"] for c in groups["admin"]] == ["TC003"]
        assert [c["id"] for c in groups["other"]] == ["TC004"]

    def test_export_split_writes_bucket_files(self, tmp_path: Path):
        cases = [
            {"module": "我要摇人-转工单", "id": "TC001", "title": "a", "type": "positive",
             "method": "GET", "path": "/x", "steps": []},
            {"module": "用户管理", "id": "TC002", "title": "b", "type": "negative",
             "method": "GET", "path": "/y", "steps": []},
        ]
        counts = export_cases_to_markdown_split(cases, tmp_path / "out")
        assert counts.get("cases.md") == 2
        assert counts.get("cases-call.md") == 1
        assert counts.get("cases-admin.md") == 1
        assert (tmp_path / "out" / "cases-call.md").exists()
        assert (tmp_path / "out" / "cases-admin.md").exists()
        assert not (tmp_path / "out" / "cases-tasks.md").exists()
