"""CLI: 向 api-test-cases.xlsx 追加数据驱动用例（按模块）。

用法:
    python automation/scripts/cli-append-cases.py --module call
    python automation/scripts/cli-append-cases.py --module admin

用例定义在本文件 CASES 字典中，格式与 Excel 列一致。
"""

import json
import sys
from pathlib import Path

import openpyxl

XLSX = Path(__file__).resolve().parents[1] / "testdata" / "cases" / "api-test-cases.xlsx"

CASES = {
    "call": [
        {"id": "CALL-026", "method": "PUT", "path": "/api/call/conversations/1", "auth": "Y",
         "payload": {"title": "Updated conv"}, "expected_status": 200,
         "expected_fields": {"id": 1, "title": "Updated conv"}, "note": "正常流程：更新会话"},
        {"id": "CALL-027", "method": "PUT", "path": "/api/call/conversations/99999", "auth": "Y",
         "payload": {"title": "x"}, "expected_status": 404, "note": "异常流程：会话不存在"},
        {"id": "CALL-028", "method": "PUT", "path": "/api/call/conversations/1", "auth": "Y",
         "payload": {"title": ""}, "expected_status": 422, "note": "数据校验：title 为空"},
        {"id": "CALL-029", "method": "DELETE", "path": "/api/call/conversations/1", "auth": "Y",
         "expected_status": 204, "note": "正常流程：删除会话"},
        {"id": "CALL-030", "method": "DELETE", "path": "/api/call/conversations/99999", "auth": "Y",
         "expected_status": 404, "note": "异常流程：会话不存在"},
        {"id": "CALL-031", "method": "GET", "path": "/api/call/messages?conversation_id=1", "auth": "Y",
         "expected_status": 200, "expected_fields": {"total": 1}, "note": "正常流程：消息列表"},
        {"id": "CALL-032", "method": "GET", "path": "/api/call/messages", "auth": "Y",
         "expected_status": 422, "note": "数据校验：缺 conversation_id"},
        {"id": "CALL-033", "method": "GET", "path": "/api/call/messages/1", "auth": "Y",
         "expected_status": 200, "expected_fields": {"id": 1}, "note": "正常流程：消息详情"},
        {"id": "CALL-034", "method": "GET", "path": "/api/call/messages/99999", "auth": "Y",
         "expected_status": 404, "note": "异常流程：消息不存在"},
        {"id": "CALL-035", "method": "PUT", "path": "/api/call/messages/1", "auth": "Y",
         "payload": {"content": "Updated"}, "expected_status": 200,
         "expected_fields": {"id": 1, "content": "Updated"}, "note": "正常流程：更新消息"},
        {"id": "CALL-036", "method": "PUT", "path": "/api/call/messages/99999", "auth": "Y",
         "payload": {"content": "x"}, "expected_status": 404, "note": "异常流程：消息不存在"},
        {"id": "CALL-037", "method": "DELETE", "path": "/api/call/messages/1", "auth": "Y",
         "expected_status": 204, "note": "正常流程：删除消息"},
        {"id": "CALL-038", "method": "DELETE", "path": "/api/call/messages/99999", "auth": "Y",
         "expected_status": 404, "note": "异常流程：消息不存在"},
        {"id": "CALL-039", "method": "GET", "path": "/api/call/my-tasks/1", "auth": "Y",
         "expected_status": 200, "expected_fields": {"id": 1}, "note": "正常流程：我的任务详情"},
        {"id": "CALL-040", "method": "POST", "path": "/api/ai/qa/submit", "auth": "N",
         "payload": {"conversation_id": 1}, "expected_status": 401, "note": "权限：未认证提交转工单"},
        {"id": "CALL-041", "auth": "Y", "expected_status": 200,
         "steps": [
             {"method": "POST", "path": "/api/call/qa/ask", "payload": {"question": "help"},
              "expected_status": 200, "expected_fields": {"success": True}},
             {"method": "POST", "path": "/api/ai/qa/submit", "payload": {"conversation_id": 1},
              "expected_status": 200, "expected_fields": {"status": "created"}},
             {"method": "POST", "path": "/api/ai/qa/ticket/ack", "payload": {"ticket_id": 1},
              "expected_status": 200, "expected_fields": {"status": "acknowledged"}},
         ], "note": "全链路：提问→转工单→确认"},
    ],
    "admin": [
        {"id": "ADMIN-025", "method": "POST", "path": "/api/admin/users", "auth": "Y",
         "payload": {"username": "newuser", "password": "pass", "name": "New User", "role": "engineer"},
         "expected_status": 201, "expected_fields": {"username": "newuser"}, "note": "正常流程：创建用户"},
        {"id": "ADMIN-026", "method": "POST", "path": "/api/admin/users", "auth": "Y",
         "payload": {}, "expected_status": 422, "note": "数据校验：缺 username"},
        {"id": "ADMIN-027", "method": "POST", "path": "/api/admin/users", "auth": "Y",
         "payload": {"username": "testadmin"}, "expected_status": 409, "note": "异常流程：用户名已存在"},
        {"id": "ADMIN-028", "method": "PUT", "path": "/api/admin/users/testadmin", "auth": "Y",
         "payload": {"name": "Admin Renamed"}, "expected_status": 200,
         "expected_fields": {"username": "testadmin"}, "note": "正常流程：更新用户"},
        {"id": "ADMIN-029", "method": "PUT", "path": "/api/admin/users/nobody", "auth": "Y",
         "payload": {"name": "x"}, "expected_status": 404, "note": "异常流程：用户不存在"},
        {"id": "ADMIN-030", "method": "POST", "path": "/api/admin/roles", "auth": "Y",
         "payload": {"name": "viewer"}, "expected_status": 201,
         "expected_fields": {"name": "viewer"}, "note": "正常流程：创建角色"},
        {"id": "ADMIN-031", "method": "POST", "path": "/api/admin/roles", "auth": "Y",
         "payload": {}, "expected_status": 422, "note": "数据校验：缺 name"},
        {"id": "ADMIN-032", "method": "POST", "path": "/api/admin/roles", "auth": "Y",
         "payload": {"name": "admin"}, "expected_status": 409, "note": "异常流程：角色已存在"},
        {"id": "ADMIN-033", "method": "PUT", "path": "/api/admin/roles/1", "auth": "Y",
         "payload": {"name": "superadmin"}, "expected_status": 200,
         "expected_fields": {"name": "superadmin"}, "note": "正常流程：更新角色"},
        {"id": "ADMIN-034", "method": "PUT", "path": "/api/admin/roles/999", "auth": "Y",
         "payload": {"name": "x"}, "expected_status": 404, "note": "异常流程：角色不存在"},
        {"id": "ADMIN-035", "method": "PUT", "path": "/api/admin/projects/1", "auth": "Y",
         "payload": {"name": "Updated Project"}, "expected_status": 200,
         "expected_fields": {"id": 1, "name": "Updated Project"}, "note": "正常流程：更新项目"},
        {"id": "ADMIN-036", "method": "PUT", "path": "/api/admin/projects/999", "auth": "Y",
         "payload": {"name": "x"}, "expected_status": 404, "note": "异常流程：项目不存在"},
        {"id": "ADMIN-037", "method": "DELETE", "path": "/api/admin/projects/1", "auth": "Y",
         "expected_status": 204, "note": "正常流程：删除项目"},
        {"id": "ADMIN-038", "method": "DELETE", "path": "/api/admin/projects/999", "auth": "Y",
         "expected_status": 404, "note": "异常流程：项目不存在"},
        {"id": "ADMIN-039", "method": "POST", "path": "/api/admin/projects/risks", "auth": "Y",
         "payload": {"name": "Risk A", "level": "high"}, "expected_status": 200,
         "expected_fields": {"name": "Risk A"}, "note": "正常流程：创建风险"},
        {"id": "ADMIN-040", "method": "POST", "path": "/api/admin/projects/risks", "auth": "Y",
         "payload": {}, "expected_status": 422, "note": "数据校验：缺 name"},
        {"id": "ADMIN-041", "method": "PUT", "path": "/api/admin/projects/risks/R1", "auth": "Y",
         "payload": {"level": "low"}, "expected_status": 200,
         "expected_fields": {"risk_code": "R1", "level": "low"}, "note": "正常流程：更新风险"},
        {"id": "ADMIN-042", "method": "PUT", "path": "/api/admin/projects/risks/R999", "auth": "Y",
         "payload": {"level": "low"}, "expected_status": 404, "note": "异常流程：风险不存在"},
        {"id": "ADMIN-043", "method": "DELETE", "path": "/api/admin/projects/risks/R1", "auth": "Y",
         "expected_status": 204, "note": "正常流程：删除风险"},
        {"id": "ADMIN-044", "method": "DELETE", "path": "/api/admin/projects/risks/R999", "auth": "Y",
         "expected_status": 404, "note": "异常流程：风险不存在"},
        {"id": "ADMIN-045", "method": "POST", "path": "/api/admin/users", "auth": "N",
         "payload": {"username": "x"}, "expected_status": 401, "note": "权限：未认证创建用户"},
    ],
}


def main() -> int:
    module = sys.argv[sys.argv.index("--module") + 1] if "--module" in sys.argv else ""
    if module not in CASES:
        print(f"Unknown module: {module!r}. Available: {sorted(CASES)}")
        return 1

    wb = openpyxl.load_workbook(XLSX)
    if module not in wb.sheetnames:
        print(f"Sheet not found: {module}")
        return 1
    ws = wb[module]
    header = [str(c.value) if c.value is not None else "" for c in ws[1]]
    existing = {str(row[0]).strip() for row in ws.iter_rows(min_row=2, values_only=True) if row[0]}

    added = 0
    for case in CASES[module]:
        cid = case["id"]
        if cid in existing:
            print(f"skip (exists): {cid}")
            continue
        row = {
            "id": cid,
            "module": module,
            "method": case.get("method", ""),
            "path": case.get("path", ""),
            "auth": case.get("auth", "Y"),
            "role": case.get("role", ""),
            "payload": json.dumps(case["payload"], ensure_ascii=False) if "payload" in case else "",
            "expected_status": case["expected_status"],
            "expected_fields": json.dumps(case["expected_fields"], ensure_ascii=False) if "expected_fields" in case else "",
            "type": case.get("type", ""),
            "note": case.get("note", ""),
            "steps": json.dumps(case["steps"], ensure_ascii=False) if "steps" in case else "",
        }
        ws.append([row.get(h, "") for h in header])
        added += 1
        print(f"added: {cid}")
    wb.save(XLSX)
    print(f"Done. Added {added} cases to sheet '{module}'.")
    return 0


if __name__ == "__main__":
    sys.exit(main())
