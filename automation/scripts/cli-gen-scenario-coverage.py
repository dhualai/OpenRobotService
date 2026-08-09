"""CLI: 生成产品场景 → 测试用例覆盖映射表。

用法:
    python automation/scripts/cli-gen-scenario-coverage.py

数据源: automation/references/ai-service-test-scenarios.xlsx（产品场景清单）
输出:   automation/docs/testing/scenario-coverage.md

映射规则维护在 _MAPPING（场景ID -> 测试函数名）与 _STATUS（场景ID -> 覆盖状态+类别）。
"""

import sys
from pathlib import Path

import openpyxl

XLSX = Path(__file__).resolve().parents[1] / "references" / "ai-service-test-scenarios.xlsx"
OUT = Path(__file__).resolve().parents[1] / "docs" / "testing" / "scenario-coverage.md"

# 场景ID -> 覆盖状态。已覆盖: (状态, 对应用例, 说明)；未覆盖: ("未覆盖", 类别, 说明)
_MAPPING = {
    # 用户端-AI在线诊断
    "U001": ("已覆盖", "CALL-005 test_ask_question", "纯文本提问"),
    "U002": ("已覆盖", "CALL-005 test_ask_question", "服务号相关提问"),
    "U003": ("未覆盖", "AI 语义", "非业务闲聊拦截，AI 评测可扩展"),
    "U004": ("未覆盖", "API(依赖AI)", "附件上传 /upload，依赖 AI 服务"),
    "U005": ("未覆盖", "API(依赖AI)", "附件上传 /upload"),
    "U006": ("未覆盖", "API(依赖AI)", "附件上传 /upload"),
    "U007": ("未覆盖", "纯UI", "消息重试，前端交互"),
    "U008": ("未覆盖", "纯UI", "剪贴板，前端交互"),
    "U009": ("已覆盖", "CALL-002/003 test_list_conversations", "历史会话持久化(API 侧)"),
    "U010": ("已覆盖", "CALL-041 test_full_flow_ask_submit_ack", "引导创建工单闭环"),
    "U011": ("已覆盖", "CALL-006 test_ask_empty_question", "空消息 422"),
    # 用户端-AI引导转工单
    "U101": ("未覆盖", "纯UI", "弹窗预填充，前端交互"),
    "U102": ("未覆盖", "纯UI", "人工编辑，前端交互"),
    "U103": ("未覆盖", "纯UI", "项目绑定下拉，前端交互"),
    "U104": ("未覆盖", "纯UI", "必填拦截，前端交互"),
    "U105": ("已覆盖", "TASK-001 test_create_task_basic", "工单提交成功(API 侧)"),
    "U106": ("未覆盖", "业务规则", "自动派单依赖真实分工配置"),
    "U107": ("未覆盖", "业务规则", "按负责人派单，依赖真实分工"),
    "U108": ("未覆盖", "业务规则", "按模块派单，依赖真实分工"),
    "U109": ("未覆盖", "业务规则", "行业分工规则，依赖配置"),
    "U110": ("未覆盖", "业务规则", "账号禁用兜底，依赖真实流程"),
    "U111": ("未覆盖", "业务规则", "无匹配兜底，依赖真实流程"),
    # 用户端-历史工单
    "U201": ("已覆盖", "TASK-004 test_list_tasks", "工单列表"),
    "U202": ("已覆盖", "TASK-013 test_filter_keyword", "模糊搜索"),
    "U203": ("已覆盖", "TASK-013/014", "状态筛选(API 侧)"),
    "U204": ("未覆盖", "纯UI", "滚动加载，前端交互"),
    "U205": ("已覆盖", "TASK-006 test_get_task_detail", "工单详情"),
    "U206": ("已覆盖", "TASK-015 test_comment_create", "发布评论"),
    "U207": ("已覆盖", "TASK-015 test_comment_create", "查看评论"),
    "U208": ("已覆盖", "CALL-015 test_cuiban", "催办"),
    "U209": ("未覆盖", "无后端API", "上报接口未实现"),
    "U210": ("未覆盖", "无后端API", "撤回接口未实现"),
    "U211": ("已覆盖", "CALL-009 test_my_tasks_list", "仅本人工单(API 侧)"),
    # 用户端-历史会话
    "HS-P0-001": ("已覆盖", "CALL-002 test_list_conversations", "会话列表"),
    "HS-P0-002": ("已覆盖", "CALL-001 test_create_conversation", "新建会话"),
    "HS-P0-003": ("已覆盖", "CALL-003 test_get_conversation_detail", "会话详情"),
    "HS-P0-004": ("已覆盖", "CALL-026 test_update_conversation", "修改标题"),
    "HS-P0-005": ("未覆盖", "纯UI", "取消修改，前端交互"),
    "HS-P0-006": ("已覆盖", "CALL-028 test_update_conversation_empty_title", "空标题 422"),
    "HS-P0-007": ("已覆盖", "CALL-029 test_delete_conversation", "删除会话"),
    "HS-P0-008": ("未覆盖", "纯UI", "取消删除，前端交互"),
    "HS-P0-009": ("未覆盖", "纯UI", "滚动加载，前端交互"),
    "HS-P0-010": ("未覆盖", "API", "会话搜索接口未实现"),
    "HS-P0-011": ("未覆盖", "API", "仅本人会话隔离"),
    "HS-P0-012": ("未覆盖", "纯UI", "时间格式，前端展示"),
    # 管理后台-系统任务
    "B001": ("已覆盖", "TASK-021/22/23 test_create_task_type_*", "多类型工单"),
    "B002": ("已覆盖", "TASK-013/014", "搜索排序筛选"),
    "B003": ("已覆盖", "TASK-006 test_get_task_detail", "工单详情"),
    "B004": ("已覆盖", "TASK-015 test_comment_create", "讨论区"),
    "B005": ("已覆盖", "TASK-008 test_update_task", "修改工单"),
    "B006": ("未覆盖", "无后端API", "退回/挂起为同步状态，无独立接口"),
    "B007": ("已覆盖", "TASK-011 test_assign_engineer", "重新指派"),
    "B008": ("未覆盖", "无后端API", "升级上报接口未实现"),
    "B009": ("未覆盖", "纯UI", "专属详情面板，前端渲染"),
    # 管理后台-跨项目看板
    "B101": ("已覆盖", "test_dashboard_tickets", "工单状态看板"),
    "B102": ("已覆盖", "test_dashboard_projects", "调度项目看板(API 侧)"),
    "B103": ("未覆盖", "API", "手动同步接口"),
    "B104": ("已覆盖", "test_dashboard_projects_urgency", "紧急度看板"),
    # 管理后台-更多
    "B201": ("已覆盖", "ADMIN-004 + test_projects_create_missing_name", "新建项目 + 必填校验"),
    "B202": ("未覆盖", "纯UI", "项目类型区分"),
    "B203": ("未覆盖", "无后端API", "授权记录接口未实现"),
    "B204": ("未覆盖", "无后端API", "授权记录接口未实现"),
    "B205": ("未覆盖", "纯UI", "机器码格式，前端校验"),
    "B206": ("已覆盖", "test_user_roles_assign", "分配角色"),
    "B207": ("未覆盖", "纯UI", "无角色提示"),
    "B301": ("未覆盖", "API(依赖文件)", "data/upload-file，文件导入"),
    "B302": ("未覆盖", "API(依赖文件)", "JSON 导入"),
    "B303": ("已覆盖", "test_daily_report", "日报生成(API 侧)"),
    "B304": ("未覆盖", "API", "周报聚合接口"),
    "B305": ("未覆盖", "API", "日报权限过滤"),
    "UM-P0-001": ("已覆盖", "ADMIN-026 test_users_create_missing_username", "必填校验"),
    "UM-P0-002": ("已覆盖", "ADMIN-025 test_users_create", "新增用户"),
    "UM-P0-003": ("已覆盖", "ADMIN-028 test_users_update", "编辑用户"),
    "UM-P0-004": ("已覆盖", "ADMIN-020 test_users_delete_cascade", "删除用户"),
    "UM-P1-001": ("未覆盖", "纯UI", "编辑按钮置灰"),
    "UM-P1-002": ("未覆盖", "纯UI", "删除按钮置灰"),
    "RM-P0-001": ("已覆盖", "ADMIN-031 test_roles_create_missing_name", "必填校验"),
    "RM-P0-002": ("已覆盖", "ADMIN-030 test_roles_create", "新增角色"),
    "RM-P0-003": ("已覆盖", "ADMIN-033 test_roles_update", "编辑角色"),
    "RM-P0-004": ("已覆盖", "test_roles_delete", "删除角色"),
    "RM-P1-001": ("未覆盖", "纯UI", "内置角色不可编辑"),
    "RM-P1-002": ("已覆盖", "test_roles_delete_builtin", "删除已分配角色 400"),
    "PM-P0-001": ("已覆盖", "test_permissions_create_missing_code", "必填校验"),
    "PM-P0-002": ("已覆盖", "test_permissions_create_ok", "新增权限"),
    "PM-P0-003": ("已覆盖", "test_permissions_update", "编辑权限"),
    "PM-P0-004": ("已覆盖", "test_permissions_delete", "删除权限"),
    "PM-P1-001": ("未覆盖", "API", "权限开关启停"),
    "PM-P1-002": ("未覆盖", "纯UI", "内置权限不可编辑"),
    "RA-P0-001": ("已覆盖", "test_user_roles_assign", "分配角色"),
    "RA-P0-002": ("已覆盖", "test_user_roles_remove", "取消角色"),
    "RA-P1-001": ("未覆盖", "API", "批量分配角色"),
    "RA-P1-002": ("未覆盖", "纯UI", "低权限隐藏选项"),
    "OL-P0-001": ("未覆盖", "无后端API", "操作记录查询接口未实现"),
    "OL-P0-002": ("未覆盖", "无后端API", "操作记录筛选"),
    "OL-P0-003": ("未覆盖", "无后端API", "操作记录详情"),
    "OL-P1-001": ("未覆盖", "无后端API", "敏感操作日志"),
    "OL-P1-002": ("未覆盖", "无后端API", "日志权限"),
    "PL-P0-001": ("已覆盖", "ADMIN-045 + permissions 矩阵", "权限链生效(API 侧)"),
    "PL-P0-002": ("未覆盖", "API", "回收权限验证"),
    "PL-P0-003": ("未覆盖", "API", "多角色叠加"),
    "PL-P0-004": ("已覆盖", "ADMIN-016/017 认证矩阵", "超级管理员(API 侧)"),
    "PL-P1-001": ("未覆盖", "API", "权限覆盖"),
    "PL-P1-002": ("未覆盖", "API", "权限缓存刷新"),
    "PC-P0-001": ("未覆盖", "纯UI", "头像点击"),
    "PC-P0-002": ("未覆盖", "纯UI", "未读角标"),
    "PC-P0-003": ("未覆盖", "纯UI", "角标消失"),
    "PC-P0-004": ("未覆盖", "纯UI", "角标一致性"),
    "PC-P0-005": ("未覆盖", "纯UI", "头像渲染"),
    "PC-P0-006": ("未覆盖", "纯UI", "弹窗渲染"),
    "PC-P0-007": ("未覆盖", "纯UI", "遮罩关闭"),
    "PC-P0-008": ("未覆盖", "纯UI", "换头像"),
    "PC-P0-009": ("未覆盖", "纯UI", "昵称编辑"),
    "PC-P0-010": ("未覆盖", "纯UI", "退出登录确认"),
    # 端到端
    "E001": ("已覆盖", "CALL-005 test_ask_question", "AI 直接解答闭环(API 侧)"),
    "E002": ("已覆盖", "CALL-041 test_full_flow_ask_submit_ack", "AI→人工派单闭环"),
    "E003": ("已覆盖", "TASK-032 test_full_flow_create_to_closed", "处理完成闭环(API 侧)"),
    "E004": ("未覆盖", "无后端API", "上报争议流程"),
    "E005": ("已覆盖", "TASK-019 test_assign_admin_transfer", "转交他人(API 侧)"),
    "E006": ("未覆盖", "AI 语义", "知识库补充后复用"),
    # AI Benchmark
    "A001": ("已覆盖", "RAG-002/004 test_recall", "调度故障精准召回"),
    "A002": ("未覆盖", "AI 语义", "拒绝回答评测可扩展"),
    "A003": ("未覆盖", "AI 语义", "同义问句评测可扩展"),
    "A004": ("未覆盖", "AI 语义", "多文档整合评测可扩展"),
    "A005": ("未覆盖", "AI 语义", "引导补充评测可扩展"),
}

_STATUS_ORDER = ["已覆盖", "未覆盖"]


def main() -> int:
    wb = openpyxl.load_workbook(XLSX, data_only=True)
    rows = []
    for ws in wb.worksheets:
        header = [str(c.value) if c.value else "" for c in ws[1]]
        for r in ws.iter_rows(min_row=2, values_only=True):
            d = dict(zip(header, r))
            cid = str(d.get("用例ID") or "").strip()
            if not cid or cid == "None":
                continue
            scene = str(d.get("测试场景") or "")
            module = str(d.get("所属模块") or "")
            priority = str(d.get("优先级") or "")
            status, detail, note = _MAPPING.get(cid, ("未覆盖", "未知", ""))
            rows.append({"id": cid, "module": module, "scene": scene,
                         "priority": priority, "status": status, "detail": detail, "note": note})

    covered = sum(1 for r in rows if r["status"] == "已覆盖")
    lines = [
        "# 产品场景 → 测试用例覆盖映射",
        "",
        f"> 数据源：`automation/references/ai-service-test-scenarios.xlsx`（产品场景清单，"
        f"由 `automation/scripts/cli-gen-scenario-coverage.py` 生成）。共 {len(rows)} 条场景，"
        f"**已覆盖 {covered} 条（{covered * 100 // len(rows)}%）**。",
        "",
        "| 状态 | 场景ID | 模块 | 测试场景 | 优先级 | 对应用例/说明 |",
        "|------|--------|------|----------|--------|----------------|",
    ]
    for r in sorted(rows, key=lambda x: (x["status"] != "已覆盖", x["id"])):
        icon = "✅" if r["status"] == "已覆盖" else "⏳"
        lines.append(f"| {icon} {r['status']} | {r['id']} | {r['module']} | {r['scene']} | "
                     f"{r['priority']} | {r['detail'] or r['note']} |")
    lines.append("")
    OUT.write_text("\n".join(lines), encoding="utf-8")
    print(f"generated: {OUT} ({len(rows)} scenarios, {covered} covered)")
    return 0


if __name__ == "__main__":
    sys.exit(main())
