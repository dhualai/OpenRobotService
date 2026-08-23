#!/usr/bin/env python3
"""Read YAML test cases → append to api-test-cases.xlsx.

Usage:
    python scripts/cli-import-cases.py path/to/new-cases.yaml
    python scripts/cli-import-cases.py testdata/templates/cases-reference.yaml --dry-run
"""

import json
import re
import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parents[2]))

import yaml
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

from automation.config.paths import CASES_FILE

EXCEL_PATH = CASES_FILE
HEADERS = ["id", "module", "method", "path", "auth", "role", "payload",
           "expected_status", "expected_fields", "type", "note", "steps"]
HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HEADER_FONT = Font(bold=True, size=10, color="FFFFFF")
THIN = Border(left=Side(style="thin"), right=Side(style="thin"),
              top=Side(style="thin"), bottom=Side(style="thin"))
REQUIRED = {"module", "method", "path", "expected_status"}


def _module_prefix(module: str) -> str:
    mapping = {
        "call": "CALL",
        "tasks": "TASK",
        "admin": "ADMIN",
        "auth": "AUTH",
    }
    return mapping.get(module, module.upper()[:4])


def _next_id(ws, module: str) -> str:
    prefix = _module_prefix(module)
    max_num = 0
    for row in ws.iter_rows(min_row=2, max_col=1, values_only=True):
        val = row[0]
        if val and isinstance(val, str) and val.startswith(prefix):
            m = re.search(r"(\d+)$", val)
            if m:
                num = int(m.group(1))
                if num > max_num:
                    max_num = num
    return f"{prefix}-{max_num + 1:03d}"


def load_yaml(path: str) -> list:
    with open(path, encoding="utf-8") as f:
        data = yaml.safe_load(f)
    if not isinstance(data, list):
        raise ValueError(f"YAML must be a list of cases, got {type(data).__name__}")
    return data


def validate(cases: list):
    errors = []
    for i, case in enumerate(cases, 1):
        missing = REQUIRED - set(case.keys())
        if missing:
            errors.append(f"Case #{i}: missing fields {missing}")
        mod = case.get("module")
        if mod and mod not in ("call", "tasks", "admin", "auth"):
            errors.append(f"Case #{i}: unknown module {mod!r}")
        method = case.get("method")
        if method and method.upper() not in ("GET", "POST", "PUT", "PATCH", "DELETE"):
            errors.append(f"Case #{i}: invalid method {method!r}")
        status = case.get("expected_status")
        if status is not None and not isinstance(status, int):
            errors.append(f"Case #{i}: expected_status must be int, got {type(status).__name__}")
    if errors:
        for e in errors:
            print(f"  ERROR: {e}")
        sys.exit(1)


def _json_dumps(val) -> str:
    if val is None:
        return ""
    if isinstance(val, str):
        return val
    return json.dumps(val, ensure_ascii=False)


def _json_loads(val):
    if isinstance(val, str):
        try:
            return json.loads(val)
        except (json.JSONDecodeError, TypeError):
            return val
    return val


def generate(excel_path: Path, cases: list, dry_run: bool = False):
    wb = openpyxl.load_workbook(excel_path)

    existing_ids = set()
    for name in wb.sheetnames:
        ws = wb[name]
        for row in ws.iter_rows(min_row=2, max_col=1, values_only=True):
            if row[0]:
                existing_ids.add(str(row[0]))

    stats = {}
    skipped = 0
    for case in cases:
        cid = case.get("id")
        if cid and cid in existing_ids:
            print(f"  {cid:>12} → SKIP (already exists)")
            skipped += 1
            continue

        mod = case["module"]
        if mod not in wb.sheetnames:
            ws = wb.create_sheet(title=mod)
            for ci, h in enumerate(HEADERS, 1):
                c = ws.cell(row=1, column=ci, value=h)
                c.font = HEADER_FONT
                c.fill = HEADER_FILL
                c.alignment = Alignment(horizontal="center")
                c.border = THIN
            ws.auto_filter.ref = f"A1:J1"
        else:
            ws = wb[mod]

        if not cid:
            cid = _next_id(ws, mod)
            case["id"] = cid
        existing_ids.add(cid)

        row_num = ws.max_row + 1
        for ci, h in enumerate(HEADERS, 1):
            val = case.get(h)
            if h in ("payload", "expected_fields"):
                val = _json_dumps(val) if val else ""
            elif h == "auth":
                val = str(val).upper() if val else "N"
            elif val is None:
                val = ""
            ws.cell(row=row_num, column=ci, value=val).border = THIN

        stats[mod] = stats.get(mod, 0) + 1
        print(f"  {case['id']:>12} → {mod} sheet")

    if skipped and not stats:
        print("\nNothing to write — all cases already exist.")
        return

    if not dry_run:
        wb.save(excel_path)
        print(f"\nSaved to {excel_path}")
    else:
        print(f"\nDry-run mode: {excel_path} not modified")

    for mod, cnt in stats.items():
        print(f"  {mod}: +{cnt}")
    total = sum(stats.values())
    print(f"  Total: +{total} case{'s' if total > 1 else ''}")


def main():
    if len(sys.argv) < 2:
        print(__doc__)
        sys.exit(1)

    yaml_path = sys.argv[1]
    dry_run = "--dry-run" in sys.argv

    print(f"Loading cases from {yaml_path}")
    cases = load_yaml(yaml_path)
    print(f"  Found {len(cases)} case{'s' if len(cases) > 1 else ''}")

    validate(cases)
    print("  Validation passed")

    generate(EXCEL_PATH, cases, dry_run=dry_run)


if __name__ == "__main__":
    main()
