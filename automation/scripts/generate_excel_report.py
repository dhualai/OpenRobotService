#!/usr/bin/env python3
"""Generate Excel test case reports from Markdown inventory files.

Reads docs/testing/test-case-inventory-*.md and creates .xlsx files.
Output: output/test-reports/{module}-test-report.xlsx + combined

Usage:
    python automation/scripts/generate_excel_report.py
"""

import re
import sys
from pathlib import Path

try:
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
except ImportError:
    print("Error: openpyxl is required. Install: pip install openpyxl")
    sys.exit(1)

DOCS_DIR = Path("docs/testing")
OUTPUT_DIR = Path("output/test-reports")

HEADER_FONT = Font(bold=True, size=11, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
CELL_FONT = Font(size=10)
THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)


def parse_tables(filepath: Path):
    """Parse Markdown tables from file. Returns list of (section_name, headers, rows)."""
    text = filepath.read_text(encoding="utf-8")
    lines = text.split("\n")
    results = []
    current_section = filepath.stem
    headers = []
    rows = []
    in_table = False

    for line in lines:
        # Track section heading
        heading = re.match(r"^#{2,3}\s+(.+)", line)
        if heading:
            current_section = heading.group(1).strip()

        if not line.startswith("|"):
            if in_table and rows:
                results.append((current_section, headers, rows))
            headers, rows = [], []
            in_table = False
            continue

        cells = [c.strip() for c in line.split("|")[1:-1]]
        if not cells:
            continue

        # Skip separator line
        if re.match(r"^[\s\-:]+$", cells[0]):
            continue

        if not in_table:
            headers = cells
            in_table = True
        else:
            row = {}
            for i, h in enumerate(headers):
                row[h] = cells[i] if i < len(cells) else ""
            rows.append(row)

    if in_table and rows:
        results.append((current_section, headers, rows))
    return results


def write_sheet(ws, title, headers, rows):
    """Write a table to an Excel worksheet."""
    ws.title = str(title)[:31]
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=col, value=h)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.alignment = Alignment(horizontal="center")
        c.border = THIN_BORDER
    for ri, row in enumerate(rows, 2):
        for ci, h in enumerate(headers, 1):
            c = ws.cell(row=ri, column=ci, value=row.get(h, ""))
            c.font = CELL_FONT
            c.border = THIN_BORDER
            c.alignment = Alignment(vertical="top")
    # Auto-fit
    for ci, h in enumerate(headers, 1):
        col_letter = openpyxl.utils.get_column_letter(ci)
        widths = [len(str(h))]
        for row in rows:
            widths.append(len(str(row.get(h, ""))))
        ws.column_dimensions[col_letter].width = min(max(widths) + 3, 55)
    # Add filter row (only if there are rows)
    if rows:
        ws.auto_filter.ref = f"A1:{openpyxl.utils.get_column_letter(len(headers))}{len(rows) + 1}"


def main():
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

    modules = [
        ("api", DOCS_DIR / "test-case-inventory-api.md"),
        ("db", DOCS_DIR / "test-case-inventory-db.md"),
        ("e2e", DOCS_DIR / "test-case-inventory-e2e.md"),
        ("ai", DOCS_DIR / "test-case-inventory-ai.md"),
        ("ui", DOCS_DIR / "test-case-inventory-ui.md"),
    ]

    combined = openpyxl.Workbook()
    combined.remove(combined.active)

    for name, path in modules:
        if not path.exists():
            print(f"  Skip: {path.name}")
            continue

        print(f"  {path.name}...", end=" ")
        tables = parse_tables(path)
        if not tables:
            print("no tables found")
            continue

        # Individual workbook
        wb = openpyxl.Workbook()
        wb.remove(wb.active)

        for section, headers, rows in tables:
            sheet_name = section.replace("#", "").strip()[:31]
            # Individual
            ws = wb.create_sheet(title=sheet_name)
            write_sheet(ws, sheet_name, headers, rows)
            # Combined
            cws = combined.create_sheet(title=f"{name}-{sheet_name}"[:31])
            write_sheet(cws, f"{name}-{sheet_name}", headers, rows)

        indv_path = OUTPUT_DIR / f"{name}-test-report.xlsx"
        wb.save(indv_path)
        sheets_count = len(tables)
        rows_count = sum(len(r) for _, _, r in tables)
        print(f"{sheets_count} sheets, {rows_count} rows -> {indv_path.name}")

    combined_path = OUTPUT_DIR / "all-modules-test-report.xlsx"
    combined.save(combined_path)
    print(f"\nCombined: {combined_path}")
    print("Done.")


if __name__ == "__main__":
    main()
