#!/usr/bin/env python3
"""Generate Excel test case reports from the cases workbook.

Reads testdata/cases/api-test-cases.xlsx and creates one report file per
sheet, plus a combined workbook.

Output: output/test-reports/{module}-test-report.xlsx + all-modules-test-report.xlsx

Usage:
    python scripts/cli-generate-report.py
"""

from pathlib import Path
import sys

sys.path.insert(0, str(Path(__file__).resolve().parents[2]))

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

from automation.config.paths import CASES_FILE

OUTPUT_DIR = Path("output/test-reports")

HEADER_FONT = Font(bold=True, size=11, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
CELL_FONT = Font(size=10)
THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)


def write_sheet(ws, headers, rows):
    """Write headers + rows into an existing worksheet."""
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=col, value=h)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.alignment = Alignment(horizontal="center")
        c.border = THIN_BORDER
    for ri, row in enumerate(rows, 2):
        for ci, val in enumerate(row, 1):
            c = ws.cell(row=ri, column=ci, value=val)
            c.font = CELL_FONT
            c.border = THIN_BORDER
            c.alignment = Alignment(vertical="top")
    for ci, h in enumerate(headers, 1):
        col_letter = openpyxl.utils.get_column_letter(ci)
        width = max([len(str(h))] + [len(str(r[ci - 1])) for r in rows]) + 3
        ws.column_dimensions[col_letter].width = min(width, 55)
    if rows:
        ws.auto_filter.ref = f"A1:{openpyxl.utils.get_column_letter(len(headers))}{len(rows) + 1}"


def main():
    if not CASES_FILE.exists():
        print(f"Cases workbook not found: {CASES_FILE}")
        return

    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

    wb = openpyxl.load_workbook(CASES_FILE, read_only=True, data_only=True)
    combined = openpyxl.Workbook()
    combined.remove(combined.active)

    for name in wb.sheetnames:
        ws = wb[name]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            print(f"  {name}: empty")
            continue
        headers = [str(h) if h is not None else "" for h in rows[0]]
        data = [r for r in rows[1:] if r[0] is not None]

        report = openpyxl.Workbook()
        write_sheet(report.active, headers, data)
        path = OUTPUT_DIR / f"{name}-test-report.xlsx"
        report.save(path)

        cws = combined.create_sheet(title=name[:31])
        write_sheet(cws, headers, data)
        print(f"  {name}: {len(data)} rows -> {path.name}")

    wb.close()

    combined_path = OUTPUT_DIR / "all-modules-test-report.xlsx"
    combined.save(combined_path)
    print(f"\nCombined: {combined_path}")
    print("Done.")


if __name__ == "__main__":
    main()
