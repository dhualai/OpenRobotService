"""Tests for cli-merge-ai-cases (AI cases -> formal Excel merge)."""

import importlib.util
import json
import shutil
import sys
from pathlib import Path

import openpyxl
import pytest

_MODULE_PATH = Path(__file__).resolve().parents[1] / "cli-merge-ai-cases.py"
_spec = importlib.util.spec_from_file_location("cli_merge_ai_cases", _MODULE_PATH)
cli = importlib.util.module_from_spec(_spec)
_spec.loader.exec_module(cli)
sys.modules["cli_merge_ai_cases"] = cli

EXCEL_PATH = cli.EXCEL_PATH

from cli_merge_ai_cases import (  # noqa: E402
    _collect_existing,
    _dedup_key,
    merge,
    mock_supported,
    module_for_path,
    plan_merge,
)

AI_ROWS = [
    {"id": "TC001", "module": "我要摇人-转工单", "method": "POST", "path": "/api/qa/submit",
     "auth": "Y", "role": "admin", "payload": "{}", "expected_status": 200,
     "expected_fields": "{}", "type": "positive", "note": "需求:REQ-11 | 转工单"},
    {"id": "TC002", "module": "系统任务-工单状态流转", "method": "POST", "path": "/api/tasks",
     "auth": "Y", "role": "admin", "payload": '{"title":"x","description":"y"}',
     "expected_status": 200, "expected_fields": "{}", "type": "flow", "note": "需求:REQ-26",
     "steps": json.dumps([
         {"method": "POST", "path": "/api/tasks", "payload": '{"title":"x","description":"y"}',
          "expected_status": 200, "expected_fields": {}},
         {"method": "PATCH", "path": "/api/tasks/{{step1.body.id}}/status",
          "payload": '{"status":"in_progress"}', "expected_status": 200, "expected_fields": {}},
     ], ensure_ascii=False)},
    {"id": "TC003", "module": "微信标签管理", "method": "POST", "path": "/api/wechat/tag",
     "auth": "Y", "role": "admin", "payload": '{"name":"vip"}', "expected_status": 200,
     "expected_fields": "{}", "type": "positive", "note": "需求:REQ-66 | 创建标签"},
    {"id": "TC004", "module": "AI对话", "method": "POST", "path": "/api/ai/task/analyze",
     "auth": "Y", "role": "admin", "payload": "{}", "expected_status": 200,
     "expected_fields": "{}", "type": "positive", "note": "需求:REQ-XX"},
    {"id": "TC005", "module": "未知接口", "method": "GET", "path": "/api/unknown/xyz",
     "auth": "N", "role": "admin", "payload": "", "expected_status": 200,
     "expected_fields": "{}", "type": "edge", "note": "边界"},
]


@pytest.fixture
def formal(tmp_path, monkeypatch):
    """Isolated copy of the formal Excel for merge tests."""
    target = tmp_path / "formal.xlsx"
    shutil.copy2(EXCEL_PATH, target)
    monkeypatch.setattr(cli, "EXCEL_PATH", target)
    return target


def _write_ai_xlsx(tmp_path: Path, rows) -> Path:
    out = tmp_path / "cases.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    headers = ["id", "module", "method", "path", "auth", "role", "payload",
               "expected_status", "expected_fields", "type", "note", "steps"]
    ws.append(headers)
    for r in rows:
        ws.append([r.get(h, "") for h in headers])
    wb.save(out)
    return out


class TestClassification:
    def test_module_by_path_prefix(self):
        assert module_for_path("/api/qa/ask", {}) == "call"
        assert module_for_path("/api/conversations/1", {}) == "call"
        assert module_for_path("/api/tasks/1/status", {}) == "tasks"
        assert module_for_path("/api/integrations", {}) == "tasks"
        assert module_for_path("/api/admin/users", {}) == "admin"
        assert module_for_path("/auth/login", {}) == "auth"
        assert module_for_path("/api/auth/login", {}) == "auth"
        assert module_for_path("/api/wechat/menu", {}) == "wechat"
        assert module_for_path("/api/ai/task/analyze", {}) == "pending"
        assert module_for_path("/api/unknown/x", {}) == "pending"

    def test_custom_module_map_overrides(self):
        custom = {"/api/ai": "ai"}
        assert module_for_path("/api/ai/task/analyze", custom) == "ai"
        assert module_for_path("/api/qa/ask", custom) == "call"

    def test_mock_supported(self):
        assert mock_supported("/api/tasks/1/status")
        assert mock_supported("/health")
        assert mock_supported("/api/wechat")
        assert mock_supported("/api/wechat/create_menu")
        assert not mock_supported("/api/wechat/tag")
        assert not mock_supported("/api/unknown/xyz")

    def test_plan_merge(self):
        plan = plan_merge(AI_ROWS, {})
        assert len(plan["call"]) == 1
        assert len(plan["tasks"]) == 1
        assert len(plan["wechat"]) == 1
        assert len(plan["pending"]) == 2


class TestMerge:
    def test_merge_appends_and_renumbers(self, formal, tmp_path):
        ai_xlsx = _write_ai_xlsx(tmp_path, AI_ROWS[:3])
        merge(ai_xlsx, dry_run=False, custom_map={}, skip_unsupported=False, backup=False)

        wb = openpyxl.load_workbook(formal)
        assert "wechat" in wb.sheetnames
        call_rows = [r for r in wb["call"].iter_rows(min_row=2, values_only=True)]
        new_call = [r for r in call_rows if "需求:REQ-11" in (r[10] or "")]
        assert len(new_call) == 1
        assert new_call[0][0].startswith("CALL-")

        wechat_rows = [r for r in wb["wechat"].iter_rows(min_row=2, values_only=True)]
        assert len(wechat_rows) == 1
        assert wechat_rows[0][0].startswith("WECHAT-")
        assert "Mock未支持" in (wechat_rows[0][10] or "")

    def test_merge_dry_run_does_not_write(self, formal, tmp_path):
        before = openpyxl.load_workbook(formal).sheetnames
        ai_xlsx = _write_ai_xlsx(tmp_path, AI_ROWS[:2])
        merge(ai_xlsx, dry_run=True, custom_map={}, skip_unsupported=False, backup=False)
        after = openpyxl.load_workbook(formal).sheetnames
        assert before == after

    def test_merge_idempotent_no_duplicates(self, formal, tmp_path):
        ai_xlsx = _write_ai_xlsx(tmp_path, AI_ROWS[:2])
        merge(ai_xlsx, dry_run=False, custom_map={}, skip_unsupported=False, backup=True)
        before = _collect_existing(openpyxl.load_workbook(formal))[1]
        merge(ai_xlsx, dry_run=False, custom_map={}, skip_unsupported=False, backup=False)
        after = _collect_existing(openpyxl.load_workbook(formal))[1]
        assert before == after

    def test_merge_skip_unsupported(self, formal, tmp_path):
        ai_xlsx = _write_ai_xlsx(tmp_path, AI_ROWS[:3])
        merge(ai_xlsx, dry_run=False, custom_map={}, skip_unsupported=True, backup=False)
        wb = openpyxl.load_workbook(formal)
        assert "wechat" not in wb.sheetnames  # /api/wechat/tag not in mock, no sheet created
        assert "call" in wb.sheetnames
        assert "tasks" in wb.sheetnames

    def test_merge_pending_not_written(self, formal, tmp_path):
        base = sum(ws.max_row - 1 for ws in openpyxl.load_workbook(formal).worksheets)
        ai_xlsx = _write_ai_xlsx(tmp_path, [AI_ROWS[3], AI_ROWS[4]])
        merge(ai_xlsx, dry_run=False, custom_map={}, skip_unsupported=False, backup=False)
        wb = openpyxl.load_workbook(formal)
        total = sum(ws.max_row - 1 for ws in wb.worksheets)
        assert total == base  # pending rows not written

    def test_merge_preserves_steps_column(self, formal, tmp_path):
        ai_xlsx = _write_ai_xlsx(tmp_path, [AI_ROWS[1]])
        merge(ai_xlsx, dry_run=False, custom_map={}, skip_unsupported=False, backup=True)
        wb = openpyxl.load_workbook(formal)
        task_rows = list(wb["tasks"].iter_rows(min_row=2, values_only=True))
        flow_row = [r for r in task_rows if "需求:REQ-26" in (r[10] or "")]
        assert len(flow_row) == 1
        steps = json.loads(flow_row[0][11])
        assert len(steps) == 2
        assert steps[1]["path"] == "/api/tasks/{{step1.body.id}}/status"

    def test_dedup_key_uses_method_path_payload(self):
        a = {"module": "tasks", "method": "POST", "path": "/api/tasks", "payload": "{}"}
        b = {"module": "tasks", "method": "POST", "path": "/api/tasks", "payload": "{}"}
        c = {"module": "tasks", "method": "POST", "path": "/api/tasks", "payload": '{"x":1}'}
        assert _dedup_key(a) == _dedup_key(b)
        assert _dedup_key(a) != _dedup_key(c)
