#!/usr/bin/env python3
"""Generate api-test-cases.xlsx from inline test case data.

One-time bootstrap. After this, update Excel directly.
"""

import json
from pathlib import Path

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

OUT = Path("automation/testdata/api/api-test-cases.xlsx")
HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HEADER_FONT = Font(bold=True, size=10, color="FFFFFF")
THIN = Border(left=Side(style="thin"), right=Side(style="thin"),
              top=Side(style="thin"), bottom=Side(style="thin"))

HEADERS = ["id", "module", "method", "path", "auth", "payload",
           "expected_status", "expected_fields", "type", "note"]

cases = {
    "tasks": [
        ("tasks-001", "tasks", "POST", "/api/tasks", "Y",
         '{"title":"Test task","description":"Test description"}', 200,
         '{"status":"pending"}', "", "Minimal required fields"),
        ("tasks-002", "tasks", "POST", "/api/tasks", "Y",
         '{"title":"Full task","description":"Full desc","priority":"high","tags":["urgent"]}', 200,
         '{"priority":"high"}', "", "Full field creation"),
        ("tasks-003", "tasks", "POST", "/api/tasks", "Y",
         '{}', 422, "", "", "Missing required fields"),
        ("tasks-004", "tasks", "GET", "/api/tasks", "Y",
         '', 200, "", "", "List tasks"),
        ("tasks-005", "tasks", "GET", "/api/tasks", "Y",
         '', 200, "", "", "Paginated list (params via query)"),
        ("tasks-006", "tasks", "GET", "/api/tasks/1", "Y",
         '', 200, '{"id":1}', "", "Task detail found"),
        ("tasks-007", "tasks", "GET", "/api/tasks/99999", "Y",
         '', 404, "", "", "Task detail not found"),
        ("tasks-008", "tasks", "PUT", "/api/tasks/1", "Y",
         '{"title":"Updated"}', 200, '{"title":"Updated"}', "", "Update task"),
        ("tasks-009", "tasks", "PATCH", "/api/tasks/1/status", "Y",
         '{"status":"in_progress"}', 200, '{"status":"in_progress"}', "", "Valid status transition"),
        ("tasks-010", "tasks", "PATCH", "/api/tasks/1/status", "Y",
         '{"status":"closed"}', 400, "", "", "Invalid status transition"),
        ("tasks-011", "tasks", "PATCH", "/api/tasks/1/assign", "Y",
         '{"assigned_to":"engineer-02"}', 200, '{"assigned_to":"engineer-02"}', "", "Assign task"),
        ("tasks-012", "tasks", "DELETE", "/api/tasks/1", "Y",
         '', 204, "", "", "Delete task"),
        ("tasks-013", "tasks", "POST", "/api/tasks/filter", "Y",
         '{"keyword":"test"}', 200, "", "", "Filter tasks"),
        ("tasks-014", "tasks", "GET", "/api/tasks/stats/overview", "Y",
         '', 200, "", "", "Task stats"),
        ("tasks-015", "tasks", "POST", "/api/tasks/1/comments", "Y",
         '{"content":"comment"}', 201, "", "", "Create comment"),
        ("tasks-016", "tasks", "GET", "/api/tasks/1/comments", "Y",
         '', 200, "", "", "List comments"),
        ("tasks-017", "tasks", "POST", "/api/tasks/99999/comments", "Y",
         '{"content":"x"}', 404, "", "", "Comment on nonexistent task"),
        ("tasks-018", "tasks", "POST", "/api/tasks/1/ai-assign", "Y",
         '', 200, "", "", "AI auto assign"),
    ],
    "wechat": [
        ("wechat-001", "wechat", "GET", "/api/wechat/health", "Y", '', 200, "", "", "Health check"),
        ("wechat-002", "wechat", "GET", "/api/wechat/get_menu", "Y", '', 200, "", "", "Get menu"),
        ("wechat-003", "wechat", "POST", "/api/wechat/create_menu", "Y",
         '{"button":[]}', 200, '{"errcode":0}', "", "Create menu"),
        ("wechat-004", "wechat", "POST", "/api/wechat/send_message", "Y",
         '{"touser":["u"],"msgtype":"text","text":{"content":"hi"}}', 200, '{"errcode":0}', "", "Send message"),
        ("wechat-005", "wechat", "GET", "/api/wechat", "Y", '', 200, "", "", "List tags"),
        ("wechat-006", "wechat", "POST", "/api/wechat", "Y",
         '{"name":"VIP"}', 200, '{"tag":{"name":"VIP"}}', "", "Create tag"),
    ],
}

OUT.parent.mkdir(parents=True, exist_ok=True)
wb = openpyxl.Workbook()
wb.remove(wb.active)

for module, rows in cases.items():
    ws = wb.create_sheet(title=module)
    for ci, h in enumerate(HEADERS, 1):
        c = ws.cell(row=1, column=ci, value=h)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.alignment = Alignment(horizontal="center")
        c.border = THIN
    for ri, row in enumerate(rows, 2):
        for ci, val in enumerate(row, 1):
            c = ws.cell(row=ri, column=ci, value=val)
            c.border = THIN
            c.alignment = Alignment(wrap_text=False, vertical="top")
    ws.auto_filter.ref = f"A1:{chr(64+len(HEADERS))}{len(rows)+1}"
    for ci in range(1, len(HEADERS) + 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(ci)].width = 18
    ws.column_dimensions["F"].width = 50  # payload
    ws.column_dimensions["H"].width = 35  # expected_fields

wb.save(OUT)
sheets = list(cases.keys())
count = sum(len(r) for r in cases.values())
print(f"Created {OUT} with {len(sheets)} sheets, {count} cases")
for s in sheets:
    print(f"  {s}: {len(cases[s])} cases")
