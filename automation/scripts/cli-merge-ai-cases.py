#!/usr/bin/env python3
"""Merge AI-generated cases into the formal Excel case library.

Reads `automation/references/generated-cases/{run_id}/cases.xlsx` (or a custom
path), normalizes rows (module by path prefix, re-numbered ids, dedup by
method+path+payload), checks Mock support, and appends to
`automation/testdata/cases/api-test-cases.xlsx`.

references/generated-cases/ is treated as READ-ONLY: this tool never writes
there. The formal library is only appended to (existing rows untouched).

Usage:
    python scripts/cli-merge-ai-cases.py --run-id demo-008 --dry-run
    python scripts/cli-merge-ai-cases.py --run-id demo-008
    python scripts/cli-merge-ai-cases.py --cases path/to/cases.xlsx --module-map map.yaml
    python scripts/cli-merge-ai-cases.py --run-id demo-008 --skip-unsupported
"""

import json
import re
import shutil
import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parents[2]))

import openpyxl
import yaml
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

from automation.config.paths import CASES_FILE

EXCEL_PATH = CASES_FILE
HEADERS = ["id", "module", "method", "path", "auth", "role", "payload",
           "expected_status", "expected_fields", "type", "note", "steps"]
HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HEADER_FONT = Font(bold=True, size=10, color="FFFFFF")
THIN = Border(left=Side(style="thin"), right=Side(style="thin"),
              top=Side(style="thin"), bottom=Side(style="thin"))

MODULE_PREFIXES = {
    "call": ("/api/call", "/api/conversations", "/api/qa", "/api/messages", "/api/my-tasks"),
    "tasks": ("/api/tasks", "/api/integrations"),
    "admin": ("/api/admin",),
    "auth": ("/api/auth", "/auth"),
    "wechat": ("/api/wechat",),
}

# Mock backend prefix support (src/mocks/backend_mock.py handle() routing).
MOCK_PREFIXES = ("/health", "/auth/login", "/auth/me", "/api/tasks",
                 "/api/admin", "/api/conversations", "/api/qa", "/api/messages",
                 "/api/my-tasks", "/api/ai", "/api/integrations")

# /api/wechat only supports these sub-routes (backend_mock._route_wechat); rest 404.
WECHAT_SUPPORTED_REST = {"", "/health", "/get_menu", "/create_menu", "/send_message"}

MODULE_IDS = {"call": "CALL", "tasks": "TASK", "admin": "ADMIN", "auth": "AUTH", "wechat": "WECHAT"}


def load_module_map(path: str) -> dict:
    """Load a custom prefix->module mapping (path prefix: module)."""
    if not path:
        return {}
    data = yaml.safe_load(Path(path).read_text(encoding="utf-8")) or {}
    if not isinstance(data, dict):
        print("  ERROR: module-map must be a mapping of path prefix -> module")
        sys.exit(1)
    return {str(k): str(v) for k, v in data.items()}


def module_for_path(path: str, custom_map: dict) -> str:
    """Classify a path into a module by prefix match."""
    for prefix, module in custom_map.items():
        if path.startswith(prefix):
            return module
    for module, prefixes in MODULE_PREFIXES.items():
        if any(path.startswith(p) for p in prefixes):
            return module
    return "pending"


def mock_supported(path: str) -> bool:
    """True when the path can be executed against the Mock backend."""
    if path.startswith("/api/wechat"):
        return path[len("/api/wechat"):] in WECHAT_SUPPORTED_REST
    for prefix in MOCK_PREFIXES:
        if path == prefix or path.startswith(prefix):
            return True
    return False


def _module_prefix(module: str) -> str:
    return MODULE_IDS.get(module, module.upper()[:4])


def _next_id(ws, module: str) -> str:
    prefix = _module_prefix(module)
    max_num = 0
    for row in ws.iter_rows(min_row=2, max_col=1, values_only=True):
        val = row[0]
        if val and isinstance(val, str) and val.startswith(prefix):
            m = re.search(r"(\d+)$", val)
            if m:
                num = int(m.group(1))
                if num > max_num:
                    max_num = num
    return f"{prefix}-{max_num + 1:03d}"


def _dedup_key(row: dict) -> tuple:
    return (row.get("module"), row.get("method"), row.get("path"), str(row.get("payload")))


def _collect_existing(wb) -> tuple:
    ids = set()
    keys = set()
    for name in wb.sheetnames:
        ws = wb[name]
        headers = [c.value for c in ws[1]] if ws.max_row >= 1 else []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row or row[0] is None:
                continue
            case = dict(zip(headers, row))
            ids.add(str(case.get("id")))
            keys.add(_dedup_key(case))
    return ids, keys


def _read_ai_cases(path: Path) -> list:
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] is None:
            continue
        rows.append(dict(zip(headers, row)))
    wb.close()
    return rows


def plan_merge(rows: list, custom_map: dict) -> dict:
    """Classify AI rows into per-module plans. Returns merge plan dict."""
    plan: dict = {"call": [], "tasks": [], "admin": [], "auth": [], "wechat": [], "pending": []}
    for row in rows:
        path = str(row.get("path") or "")
        module = module_for_path(path, custom_map)
        if module not in plan:
            module = "pending"
        plan[module].append(row)
    return plan


def merge(cases_path: Path, dry_run: bool, custom_map: dict, skip_unsupported: bool,
          backup: bool = True):
    print(f"Reading AI cases from {cases_path}")
    rows = _read_ai_cases(cases_path)
    print(f"  Found {len(rows)} row(s)")
    if not rows:
        return

    plan = plan_merge(rows, custom_map)

    if backup and not dry_run:
        bak = str(EXCEL_PATH) + ".bak"
        shutil.copy2(EXCEL_PATH, bak)
        print(f"Backup created: {bak}")

    wb = openpyxl.load_workbook(EXCEL_PATH)
    existing_ids, existing_keys = _collect_existing(wb)

    stats = {}
    for module in ("call", "tasks", "admin", "auth", "wechat"):
        module_rows = plan[module]
        if not module_rows:
            continue
        to_write = []
        deduped = 0
        skipped_unsupported = 0
        for row in module_rows:
            case = dict(row)
            case["module"] = module
            key = _dedup_key(case)
            if key in existing_keys:
                deduped += 1
                continue
            if not mock_supported(str(case.get("path") or "")):
                if skip_unsupported:
                    skipped_unsupported += 1
                    continue
                note = str(case.get("note") or "")
                case["note"] = (note + " | Mock未支持" if note else "Mock未支持")
            to_write.append(case)

        if not to_write:
            continue
        ws = wb[module] if module in wb.sheetnames else _create_sheet(wb, module)
        added = 0
        marked_unsupported = 0
        for case in to_write:
            if "Mock未支持" in str(case.get("note") or ""):
                marked_unsupported += 1
            cid = _next_id(ws, module)
            while cid in existing_ids:
                cid = _next_id(ws, module)
            case["id"] = cid
            existing_ids.add(cid)
            existing_keys.add(_dedup_key(case))

            row_num = ws.max_row + 1
            for ci, h in enumerate(HEADERS, 1):
                val = case.get(h, "")
                if h == "auth":
                    val = str(val).strip().upper()[:1] if val else "N"
                elif val is None:
                    val = ""
                elif h in ("payload", "expected_fields", "steps") and isinstance(val, dict):
                    val = json.dumps(val, ensure_ascii=False)
                ws.cell(row=row_num, column=ci, value=val).border = THIN
            added += 1

        stats[module] = {"added": added, "deduped": deduped,
                         "marked_unsupported": marked_unsupported,
                         "skipped_unsupported": skipped_unsupported}
        print(f"  {module:8} +{added}  (dup {deduped}, "
              f"mock-unsupported {marked_unsupported + skipped_unsupported})")

    pending = plan["pending"]
    if pending:
        print(f"  pending  {len(pending)} row(s) without module mapping (not written):")
        for row in pending[:10]:
            print(f"    {str(row.get('method')):6} {row.get('path')}  [{row.get('id')}]")
        if len(pending) > 10:
            print(f"    ... and {len(pending) - 10} more")

    if not any(s["added"] for s in stats.values()):
        print("\nNothing to write.")
        return

    if dry_run:
        print("\nDry-run mode: api-test-cases.xlsx NOT modified")
    else:
        wb.save(EXCEL_PATH)
        print(f"\nSaved to {EXCEL_PATH}")
        _print_next_steps(stats)


def _create_sheet(wb, module: str):
    ws = wb.create_sheet(title=module)
    for ci, h in enumerate(HEADERS, 1):
        c = ws.cell(row=1, column=ci, value=h)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.alignment = Alignment(horizontal="center")
        c.border = THIN
    return ws


def _print_next_steps(stats: dict):
    new_sheets = [m for m, s in stats.items() if s["added"] and m not in ("call", "tasks", "admin", "auth")]
    print("Next steps:")
    for m, s in stats.items():
        if s["added"]:
            print(f"  1. pytest tests/{m}/ -v           # verify {s['added']} new case(s)")
    if new_sheets:
        print(f"  2. python scripts/cli-generate-test-modules.py   # generate test file for new sheet(s): {', '.join(new_sheets)}")
    print("  3. pytest --alluredir=output/allure-results && allure generate output/allure-results -o output/allure-report --clean")


def main():
    args = sys.argv[1:]
    if not args or any(a in ("-h", "--help") for a in args):
        print(__doc__)
        sys.exit(0)
    dry_run = "--dry-run" in args
    skip_unsupported = "--skip-unsupported" in args
    run_id = None
    cases = None
    module_map = None
    if "--run-id" in args:
        run_id = args[args.index("--run-id") + 1]
    if "--cases" in args:
        cases = args[args.index("--cases") + 1]
    if "--module-map" in args:
        module_map = args[args.index("--module-map") + 1]

    if not cases:
        if not run_id:
            print("ERROR: provide --run-id <id> or --cases <path>")
            sys.exit(1)
        cases = Path("references/generated-cases") / run_id / "cases.xlsx"
    cases_path = Path(cases)
    if not cases_path.exists():
        print(f"ERROR: cases file not found: {cases_path}")
        sys.exit(1)

    custom_map = load_module_map(module_map)
    merge(cases_path, dry_run=dry_run, custom_map=custom_map, skip_unsupported=skip_unsupported)


if __name__ == "__main__":
    main()
