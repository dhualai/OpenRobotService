"""CLI: 批量修正 api-test-cases.xlsx 中的路径字段（mock/Excel 与真实后端契约对齐）。

用法:
    python automation/scripts/cli-update-cases-xlsx.py

幂等：已修正的路径不会被二次替换（替换后不再匹配旧前缀）。
"""

import sys
from pathlib import Path

import openpyxl

XLSX = Path(__file__).resolve().parents[1] / "testdata" / "cases" / "api-test-cases.xlsx"

# (旧前缀, 新前缀)，长前缀在前
PREFIX_MAP = [
    ("/api/admin/export/project/P001/project/P001", "/api/admin/export/project/P001"),  # 历史误伤修复
    ("/api/admin/resource-manager/resources", "/api/admin/resource-manager/resources"),  # 幂等锚点
    ("/api/admin/export/project/", "/api/admin/export/project/"),  # 幂等锚点
    ("/api/admin/resources", "/api/admin/resource-manager/resources"),
    ("/api/qa/ask/stream", "/api/call/qa/ask/stream"),
    ("/api/qa/ask", "/api/call/qa/ask"),
    ("/api/conversations", "/api/call/conversations"),
    ("/api/messages", "/api/call/messages"),
    ("/api/my-tasks/", "/api/call/my-tasks/"),
    ("/auth/login", "/api/auth/login"),
    ("/auth/me", "/api/auth/me"),
    ("/api/admin/export", "/api/admin/export/project/P001"),
]


def rewrite(path: str) -> str:
    for old, new in PREFIX_MAP:
        if path.startswith(old):
            return new + path[len(old):]
    return path


def main() -> int:
    wb = openpyxl.load_workbook(XLSX)
    changed = 0
    for ws in wb.worksheets:
        header = [str(c.value) if c.value is not None else "" for c in ws[1]]
        if "path" not in header:
            continue
        col = header.index("path") + 1
        for row in ws.iter_rows(min_row=2):
            cell = row[col - 1]
            raw = str(cell.value or "")
            if not raw:
                continue
            new = rewrite(raw)
            if new != raw:
                cell.value = new
                changed += 1
    wb.save(XLSX)
    print(f"Updated {changed} path cells in {XLSX.name}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
