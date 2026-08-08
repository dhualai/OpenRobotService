"""Data-driven API test cases: load Excel rows into parametrized pytest cases."""

import json
from typing import Optional

import openpyxl

from automation.config.paths import CASES_FILE

_CACHE: dict = {}


def load_cases(module_name: str, sheet_name: Optional[str] = None) -> list:
    """Load test cases for a module from Excel, cached after first load."""
    key = f"{module_name}:{sheet_name or module_name}"
    if key in _CACHE:
        return _CACHE[key]

    wb = openpyxl.load_workbook(CASES_FILE, read_only=True, data_only=True)
    ws_name = sheet_name or module_name
    if ws_name not in wb.sheetnames:
        wb.close()
        return []
    ws = wb[ws_name]

    headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    cases = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] is None:
            continue
        case = dict(zip(headers, row))
        # Parse JSON strings
        for field in ("payload", "expected_fields", "steps"):
            val = case.get(field)
            if isinstance(val, str) and val.strip():
                try:
                    case[field] = json.loads(val)
                except json.JSONDecodeError:
                    pass
            elif val is None:
                case[field] = {} if field == "payload" else None
        if case.get("steps") == "":
            case["steps"] = None
        # Normalize field names
        case["auth"] = str(case.get("auth", "N")).strip().upper()[:1]
        case["expected_status"] = int(case.get("expected_status", 200))
        cases.append(case)

    wb.close()
    _CACHE[key] = cases
    return cases
