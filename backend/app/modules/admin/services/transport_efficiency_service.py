"""搬运效率分析数据服务。

数据来源：
1. 数据导入（文件导入 .bz2/.json 数据包）→ 落库 CollectionData（GroupEfficiency 原始指标），
   本服务读取后按 ProjectMetricsList 参考逻辑计算搬运效率汇总（summary）与各组对比（robots）；
2. 数据导入（JSON 导入 / 旧 Excel 导入）→ 落库 ProjectTransportEfficiency / 型号明细表。
"""
from typing import Dict, List, Optional, Tuple
from datetime import datetime, timedelta
import io

from sqlalchemy import create_engine, and_, func
from sqlalchemy.orm import sessionmaker

from app.modules.admin.models_das.models import (
    ProjectTransportEfficiency,
    ProjectTransportEfficiencyRobot,
)
from app.modules.admin.services.data_service import DataService as AdminDataService
from app.modules.admin.utils_das.config import DATABASE_URL

engine = create_engine(DATABASE_URL)
SessionLocal = sessionmaker(autocommit=False, autoflush=False, bind=engine)

# 汇总表：中文指标名 -> 字段名
SUMMARY_FIELD_MAP = {
    "总任务数": "total_tasks",
    "搬运任务数量": "carry_task_count",
    "有效工作时长": "effective_work_hours",
    "机器人故障时长": "fault_hours",
    "空闲无任务时间": "idle_hours",
    "平均错误次数": "avg_error_count",
    "平均单次故障时间": "avg_fault_duration_minutes",
    "平均单次搬运任务时间": "avg_carry_duration_minutes",
    "平均切手动次数": "avg_manual_switch_count",
    "人工干预率": "manual_intervention_rate",
}

SUMMARY_INT_FIELDS = {"total_tasks", "carry_task_count"}

# 型号对比表：中文指标名 -> 字段名
ROBOT_FIELD_MAP = {
    "搬运任务总数(个)": "carry_task_total",
    "有效工作时长(h)": "effective_work_hours",
    "有效搬运效率(小时/个)": "effective_efficiency",
    "机器人故障时间(小时)": "fault_hours",
    "无工作时间(小时)": "idle_hours",
    "平均单次故障(分钟)": "avg_fault_duration_minutes",
    "平均单次搬运时间(分钟)": "avg_carry_duration_minutes",
}

ROBOT_INT_FIELDS = {"carry_task_total"}


class TransportEfficiencyService:

    def upsert_daily_summary(self, project_code: str, report_date: str, fields: Dict) -> Dict:
        db = SessionLocal()
        try:
            now_str = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            record = db.query(ProjectTransportEfficiency).filter(
                ProjectTransportEfficiency.project_code == project_code,
                ProjectTransportEfficiency.report_date == report_date,
            ).first()

            allowed_fields = set(SUMMARY_FIELD_MAP.values())
            clean_fields = {k: v for k, v in fields.items() if k in allowed_fields}

            if record:
                for key, value in clean_fields.items():
                    setattr(record, key, value)
                record.updated_at = now_str
            else:
                record = ProjectTransportEfficiency(
                    project_code=project_code,
                    report_date=report_date,
                    created_at=now_str,
                    updated_at=now_str,
                    **clean_fields,
                )
                db.add(record)

            db.commit()
            db.refresh(record)
            return self._summary_to_dict(record)
        finally:
            db.close()

    def upsert_robot_rows(self, project_code: str, report_date: str, rows: List[Dict]) -> List[Dict]:
        db = SessionLocal()
        try:
            now_str = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

            db.query(ProjectTransportEfficiencyRobot).filter(
                ProjectTransportEfficiencyRobot.project_code == project_code,
                ProjectTransportEfficiencyRobot.report_date == report_date,
            ).delete()

            allowed_fields = set(ROBOT_FIELD_MAP.values())
            created = []
            for row in rows:
                robot_model = row.get("robot_model")
                if not robot_model:
                    continue
                clean_row = {k: v for k, v in row.items() if k in allowed_fields}
                record = ProjectTransportEfficiencyRobot(
                    project_code=project_code,
                    report_date=report_date,
                    robot_model=robot_model,
                    created_at=now_str,
                    **clean_row,
                )
                db.add(record)
                created.append(record)

            db.commit()
            for record in created:
                db.refresh(record)
            return [self._robot_to_dict(record) for record in created]
        finally:
            db.close()

    def get_daily_efficiency(self, project_code: str, report_date: str) -> Optional[Dict]:
        db = SessionLocal()
        try:
            record = db.query(ProjectTransportEfficiency).filter(
                ProjectTransportEfficiency.project_code == project_code,
                ProjectTransportEfficiency.report_date == report_date,
            ).first()
            return self._summary_to_dict(record) if record else None
        finally:
            db.close()

    def get_robot_efficiency(self, project_code: str, report_date: str) -> List[Dict]:
        db = SessionLocal()
        try:
            records = db.query(ProjectTransportEfficiencyRobot).filter(
                ProjectTransportEfficiencyRobot.project_code == project_code,
                ProjectTransportEfficiencyRobot.report_date == report_date,
            ).all()
            return [self._robot_to_dict(record) for record in records]
        finally:
            db.close()

    def get_latest_manual_switch_count(self, project_code: str) -> Optional[float]:
        db = SessionLocal()
        try:
            record = db.query(ProjectTransportEfficiency).filter(
                ProjectTransportEfficiency.project_code == project_code,
            ).order_by(ProjectTransportEfficiency.report_date.desc()).first()
            return record.avg_manual_switch_count if record else None
        finally:
            db.close()

    def get_latest_manual_switch_counts(self, project_codes: List[str]) -> Dict[str, Optional[float]]:
        """批量获取多项目最新一条记录的切手动次数（一条自连接查询）。

        替代循环内逐项目调用 get_latest_manual_switch_count（N 条查询 → 1 条）。
        返回 {project_code: avg_manual_switch_count|None}；无数据的项目不出现在 dict 中。
        排序口径与单条版本一致：report_date 为 String 列，按字符串 desc 取最大（格式统一 YYYY-MM-DD）。
        """
        if not project_codes:
            return {}
        db = SessionLocal()
        try:
            latest = (
                db.query(
                    ProjectTransportEfficiency.project_code.label("pc"),
                    func.max(ProjectTransportEfficiency.report_date).label("md"),
                )
                .filter(ProjectTransportEfficiency.project_code.in_(project_codes))
                .group_by(ProjectTransportEfficiency.project_code)
                .subquery()
            )
            rows = (
                db.query(
                    ProjectTransportEfficiency.project_code,
                    ProjectTransportEfficiency.avg_manual_switch_count,
                )
                .join(
                    latest,
                    and_(
                        ProjectTransportEfficiency.project_code == latest.c.pc,
                        ProjectTransportEfficiency.report_date == latest.c.md,
                    ),
                )
                .all()
            )
            return {row.project_code: row.avg_manual_switch_count for row in rows}
        finally:
            db.close()

    def get_collection_summary_and_robots(self, project_code: str, date: str) -> Tuple[Optional[Dict], List[Dict], Optional[str]]:
        """从数据导入落库的 CollectionData 读取 GroupEfficiency 数据，计算搬运效率汇总。

        返回 (summary, robots, collection_time)。找不到当日数据时 summary 为 None、robots 为空。
        计算逻辑与参考页 ProjectMetricsList.jsx 的 calculateAverages / 各组数据对比一致。
        """
        # 查询窗口前后各放宽 1 天：记录的 start/end 时间戳按数据源自身时区（如 +02:00）的当日零点落库，
        # 与服务器本地时区（如 +08:00）的日期零点存在时区错位，直接按目标日查询会把记录过滤掉。
        # 宽窗口保证记录能被查到，之后再用记录自带 start_time（ISO 字符串，含原始时区）与目标日期精确匹配。
        day = datetime.strptime(date, "%Y-%m-%d")
        start = (day - timedelta(days=1)).strftime("%Y-%m-%d 00:00:00")
        end = (day + timedelta(days=2)).strftime("%Y-%m-%d 00:00:00")
        result = AdminDataService.get_collection_data_for_indicators(
            project=project_code,
            tag="GroupEfficiency",
            indicators=["*"],
            start_time=start,
            end_time=end,
        )
        content = result.get("content") if isinstance(result, dict) else None
        if not content:
            return None, [], result.get("collection_time") if isinstance(result, dict) else None

        # 宽窗口可能同时捞出相邻日期的记录（记录 end 为次日零点、且含数据源时区偏移），
        # 必须按记录自带 start_time 的日期归属严格匹配目标日；只有全部记录都缺失 start_time
        # （老数据）时才退回第一条，避免 09-03 的记录在查询 09-04 时被错误命中。
        metrics = None
        dateless_metrics = None
        saw_dated_record = False
        for data_obj in content:
            found = _find_metrics(data_obj)
            if not found:
                continue
            start_iso = data_obj.get("start_time") if isinstance(data_obj, dict) else None
            if isinstance(start_iso, str) and len(start_iso) >= 10:
                saw_dated_record = True
                if start_iso[:10] == date:
                    metrics = found
                    break
            elif dateless_metrics is None:
                dateless_metrics = found
        if metrics is None and not saw_dated_record:
            metrics = dateless_metrics
        if not metrics:
            return None, [], result.get("collection_time") if isinstance(result, dict) else None

        summary, robots = self._metrics_to_summary_and_robots(metrics)
        return summary, robots, result.get("collection_time") if isinstance(result, dict) else None

    def _metrics_to_summary_and_robots(self, metrics: Dict) -> Tuple[Dict, List[Dict]]:
        """把 GroupEfficiency 指标对象转换为 summary + robots（各组数据对比）。"""
        task_number = (metrics.get("dataIndicators") or {}).get("taskNumber") or {}

        effect_work_time = metrics.get("effectWorkTime") or {}
        robot_error_time = metrics.get("robotErrorTime") or {}
        no_work_time = metrics.get("noWorkTime") or {}
        per_error_time = metrics.get("perErrorTime") or []
        average_carry_time = metrics.get("averageCarryTime") or []

        groups = list(effect_work_time.keys())
        group_count = len(groups) or 1

        total_effect = sum((effect_work_time.get(g) or {}).get("diffTimeSeconds", 0) or 0 for g in groups)
        total_fault = sum((robot_error_time.get(g) or {}).get("totalTimeSeconds", 0) or 0 for g in groups)
        total_idle = sum((no_work_time.get(g) or {}).get("totalIdleTimeSeconds", 0) or 0 for g in groups)
        total_error_num = sum((e.get("errorNum") or 0) for e in per_error_time)
        total_per_error = sum((e.get("perErrorTimeSeconds") or 0) for e in per_error_time)
        total_per_carry = sum((c.get("perGroupSingleTaskSeconds") or 0) for c in average_carry_time)
        carry_len = len(average_carry_time) or 1

        summary = {
            "total_tasks": task_number.get("totalTasks"),
            "carry_task_count": task_number.get("carry"),
            "effective_work_hours": round(total_effect / 3600 / group_count, 2),
            "fault_hours": round(total_fault / 3600 / group_count, 2),
            "idle_hours": round(total_idle / 3600 / group_count, 2),
            "avg_error_count": round(total_error_num / group_count, 2),
            "avg_fault_duration_minutes": round(total_per_error / 60 / group_count, 2),
            "avg_carry_duration_minutes": round(total_per_carry / 60 / carry_len, 2),
            "avg_manual_switch_count": (metrics.get("averageManualCount") or {}).get("averageManualCount"),
            "manual_intervention_rate": _parse_rate((metrics.get("rateArtificialIntervention") or {}).get("rateArtificialIntervention")),
        }

        error_map = {e.get("robotGroup"): e for e in per_error_time if e.get("robotGroup")}
        carry_map = {c.get("robotGroup"): c for c in average_carry_time if c.get("robotGroup")}

        robots = []
        for group in groups:
            eff_seconds = (effect_work_time.get(group) or {}).get("diffTimeSeconds") or 0
            fault_seconds = (robot_error_time.get(group) or {}).get("totalTimeSeconds") or 0
            idle_seconds = (no_work_time.get(group) or {}).get("totalIdleTimeSeconds") or 0
            carry_info = carry_map.get(group) or {}
            error_info = error_map.get(group) or {}
            task_count = carry_info.get("taskCount") or 0
            eff_hours = eff_seconds / 3600
            robots.append({
                "robot_model": group,
                "carry_task_total": task_count,
                "effective_work_hours": round(eff_hours, 2),
                "effective_efficiency": round(eff_hours / task_count, 2) if task_count else 0,
                "fault_hours": round(fault_seconds / 3600, 2),
                "idle_hours": round(idle_seconds / 3600, 2),
                "avg_fault_duration_minutes": round((error_info.get("perErrorTimeSeconds") or 0) / 60, 2),
                "avg_carry_duration_minutes": round((carry_info.get("perGroupSingleTaskSeconds") or 0) / 60, 2),
            })

        return summary, robots


    def parse_excel(self, file_bytes: bytes) -> Tuple[Dict, List[Dict]]:
        from openpyxl import load_workbook

        wb = load_workbook(io.BytesIO(file_bytes), data_only=True)

        summary: Dict = {}
        if "汇总" in wb.sheetnames:
            ws = wb["汇总"]
            for row in ws.iter_rows(min_row=1, values_only=True):
                if not row or len(row) < 2:
                    continue
                label, value = row[0], row[1]
                if label is None:
                    continue
                field = SUMMARY_FIELD_MAP.get(str(label).strip())
                if not field or value is None:
                    continue
                summary[field] = self._coerce_number(field, value, SUMMARY_INT_FIELDS)

        robot_rows: List[Dict] = []
        if "机型明细" in wb.sheetnames:
            ws = wb["机型明细"]
            rows = list(ws.iter_rows(values_only=True))
            if rows:
                header = rows[0]
                robot_models = [str(cell).strip() if cell is not None else None for cell in header[1:]]
                robot_data: Dict[str, Dict] = {model: {"robot_model": model} for model in robot_models if model}

                for row in rows[1:]:
                    if not row:
                        continue
                    label = row[0]
                    if label is None:
                        continue
                    field = ROBOT_FIELD_MAP.get(str(label).strip())
                    if not field:
                        continue
                    for idx, model in enumerate(robot_models):
                        if not model:
                            continue
                        cell_value = row[idx + 1] if idx + 1 < len(row) else None
                        if cell_value is None:
                            continue
                        robot_data[model][field] = self._coerce_number(field, cell_value, ROBOT_INT_FIELDS)

                robot_rows = list(robot_data.values())

        return summary, robot_rows

    def _coerce_number(self, field: str, value, int_fields: set):
        try:
            num = float(value)
        except (TypeError, ValueError):
            return None
        if field in int_fields:
            return int(num)
        return num

    def _summary_to_dict(self, record: ProjectTransportEfficiency) -> Dict:
        return {
            "id": record.id,
            "project_code": record.project_code,
            "report_date": record.report_date,
            "total_tasks": record.total_tasks,
            "carry_task_count": record.carry_task_count,
            "effective_work_hours": record.effective_work_hours,
            "fault_hours": record.fault_hours,
            "idle_hours": record.idle_hours,
            "avg_error_count": record.avg_error_count,
            "avg_fault_duration_minutes": record.avg_fault_duration_minutes,
            "avg_carry_duration_minutes": record.avg_carry_duration_minutes,
            "avg_manual_switch_count": record.avg_manual_switch_count,
            "manual_intervention_rate": record.manual_intervention_rate,
            "created_at": record.created_at,
            "updated_at": record.updated_at,
        }

    def _robot_to_dict(self, record: ProjectTransportEfficiencyRobot) -> Dict:
        return {
            "id": record.id,
            "project_code": record.project_code,
            "report_date": record.report_date,
            "robot_model": record.robot_model,
            "carry_task_total": record.carry_task_total,
            "effective_work_hours": record.effective_work_hours,
            "effective_efficiency": record.effective_efficiency,
            "fault_hours": record.fault_hours,
            "idle_hours": record.idle_hours,
            "avg_fault_duration_minutes": record.avg_fault_duration_minutes,
            "avg_carry_duration_minutes": record.avg_carry_duration_minutes,
            "created_at": record.created_at,
        }


def _parse_rate(value) -> Optional[float]:
    """人工干预率解析为小数（0~1），供前端按百分比展示。

    数据源为 "10.0%" 之类百分比字符串时除以 100；已是数值则原样返回。
    """
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).strip()
    try:
        num = float(text.rstrip('%'))
    except ValueError:
        return None
    return round(num / 100, 4) if text.endswith('%') else num


def _find_metrics(obj, depth: int = 0) -> Optional[Dict]:
    """在 CollectionData 的 data JSON 中递归查找 GroupEfficiency 指标对象。

    数据包结构：{data: [{data: [metrics], ...}], start_time, end_time}，
    指标对象以 effectWorkTime 或 dataIndicators 字段为标识。
    """
    if depth > 6:
        return None
    if isinstance(obj, dict):
        if 'effectWorkTime' in obj or 'dataIndicators' in obj:
            return obj
        for value in obj.values():
            found = _find_metrics(value, depth + 1)
            if found:
                return found
    elif isinstance(obj, list):
        for value in obj:
            found = _find_metrics(value, depth + 1)
            if found:
                return found
    return None


transport_efficiency_service = TransportEfficiencyService()
